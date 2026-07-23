import os
import streamlit as st
st.write("FILES:", os.listdir())
import base64
import io
import itertools
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from urllib.parse import urlencode

import altair as alt
import duckdb
import pandas as pd
import streamlit as st


# Bộ đếm tạo key duy nhất cho các widget trong mỗi lần Streamlit chạy lại.
_WIDGET_KEY_COUNTER = itertools.count()


def unique_widget_key(prefix: str) -> str:
    return f"{prefix}_{next(_WIDGET_KEY_COUNTER)}"


def query_param_values(name: str) -> list[str]:
    """Đọc an toàn một tham số URL có thể xuất hiện nhiều lần."""
    try:
        values = st.query_params.get_all(name)
    except (AttributeError, TypeError):
        value = st.query_params.get(name)
        values = value if isinstance(value, list) else ([value] if value else [])
    return [str(value) for value in values if value not in (None, "")]


# ============================================================
# CẤU HÌNH
# ============================================================

st.set_page_config(
    page_title="BOT WO",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded",
)


st.markdown(
    """
    <style>
    :root {
        --brand:#e6002d;
        --brand-dark:#c90028;
        --text:#142033;
        --muted:#738098;
        --line:#e7ebf1;
        --page:#f7f9fc;
        --surface:#ffffff;
    }

    html, body, [class*="css"] { font-family: Inter, "Segoe UI", Arial, sans-serif; }
    .stApp { background:var(--page); color:var(--text); }
    [data-testid="stHeader"] { background:transparent; }
    [data-testid="stToolbar"] { display:none; }
    #MainMenu { visibility:hidden; }
    footer { visibility:hidden; }

    .block-container {
        max-width:none !important;
        padding:1.1rem 1.4rem 7.2rem !important;
    }

    /* Luôn hiển thị menu bên trái trên màn hình máy tính */
    [data-testid="stSidebar"] {
        display:block !important;
        visibility:visible !important;
        transform:none !important;
        background:#fff;
        border-right:1px solid var(--line);
        width:300px !important;
        min-width:300px !important;
        max-width:300px !important;
        left:0 !important;
    }

    [data-testid="stSidebarHeader"] {
        position:absolute !important;
        top:6px !important;
        left:0 !important;
        right:6px !important;
        height:36px !important;
        min-height:36px !important;
        padding:0 !important;
        z-index:50 !important;
        background:transparent !important;
        pointer-events:none !important;
    }

    [data-testid="stSidebarCollapseButton"] {
        margin-left:auto !important;
        pointer-events:auto !important;
    }

    [data-testid="stSidebar"] > div:first-child {
        display:block !important;
        visibility:visible !important;
        width:300px !important;
    }

    /* Giữ nút thu gọn/mở sidebar để người dùng chủ động điều khiển */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="collapsedControl"] {
        display:flex !important;
        visibility:visible !important;
    }
    [data-testid="stSidebar"] > div:first-child {
        position:relative !important;
        padding:0 1rem 1.5rem !important;
    }

    .sidebar-viettel-logo-wrap {
        position:relative;
        width:100%;
        height:104px;
        min-height:104px;
        display:flex;
        align-items:center;
        justify-content:center;
        box-sizing:border-box;
        padding:18px 28px 10px;
        margin:0 0 12px;
        background:#ffffff;
        border-bottom:1px solid var(--line);
        z-index:1;
        overflow:hidden;
        pointer-events:none;
    }
    .sidebar-viettel-logo {
        display:block;
        width:150px;
        max-width:78%;
        height:auto;
        margin:0 auto;
        object-fit:contain;
    }

    .sidebar-brand {
        display:flex; align-items:center; gap:10px;
        padding:0 2px 18px;
        margin-top:0;
        font-size:1.02rem; font-weight:800; color:var(--text);
    }
    .sidebar-brand-icon,.app-logo {
        display:grid; place-items:center; color:#fff;
        background:linear-gradient(135deg,var(--brand-dark),#ff334d);
        box-shadow:0 8px 20px rgba(230,0,45,.20);
    }
    .sidebar-brand-icon { width:36px; height:36px; border-radius:11px; font-size:16px; }
    .app-logo { width:38px; height:38px; border-radius:11px; font-size:17px; flex:0 0 auto; }

    .app-header {
        display:flex; align-items:center; justify-content:space-between;
        gap:16px; width:100%; box-sizing:border-box;
        background:#fff; border:1px solid var(--line); border-radius:14px;
        padding:10px 16px; margin:0 0 .8rem;
        box-shadow:0 8px 22px rgba(20,32,51,.06);
    }
    .app-header-left { display:flex; align-items:center; gap:11px; min-width:0; }
    .app-copy { min-width:0; }
    .app-header h1 { margin:0; color:#e6002d !important; font-size:2rem !important; line-height:1.15; font-weight:850; white-space:nowrap; }
    .app-header p { margin:3px 0 0; color:var(--muted); font-size:.78rem; }
    .author-line { margin-top:4px; display:flex; gap:5px; align-items:center; flex-wrap:wrap; font-size:.72rem; color:#3c4658; }
    .author-line b { color:var(--text); }
    .author-line a { color:var(--brand); text-decoration:none; font-weight:700; }
    .viettel-logo { width:96px; max-width:14vw; height:auto; object-fit:contain; flex:0 0 auto; }

    .welcome-card {
        max-width:980px; margin:3.3rem auto 1.2rem; padding:2.2rem;
        text-align:center; background:#fff; border:1px solid var(--line);
        border-radius:22px; box-shadow:0 16px 40px rgba(20,32,51,.06);
    }
    .welcome-icon {
        width:60px; height:60px; margin:0 auto 14px; display:grid; place-items:center;
        border-radius:18px; color:#fff; font-size:28px;
        background:linear-gradient(135deg,var(--brand-dark),#ff334d);
        box-shadow:0 12px 25px rgba(230,0,45,.22);
    }
    .welcome-card h2 { margin:.2rem 0 .45rem; font-size:1.45rem; color:var(--text); }
    .welcome-card p { margin:0 auto; max-width:650px; color:var(--muted); line-height:1.6; }
    .sample-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; max-width:820px; margin:1.15rem auto 0; }
    .sample-item { text-align:left; background:#fafbfe; border:1px solid var(--line); border-radius:13px; padding:12px 14px; color:#465269; font-size:.9rem; }

    /* Sidebar giống mẫu */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] {
        border:1px solid #dfe5ee !important;
        border-radius:13px !important;
        padding:10px !important;
        background:#fbfcfe !important;
        box-shadow:none !important;
    }

    [data-testid="stSidebar"] [data-testid="stFileUploader"] section {
        min-height:92px !important;
        height:92px !important;
        border:0 !important;
        background:transparent !important;
        padding:12px 8px !important;
        display:flex !important;
        align-items:center !important;
        justify-content:center !important;
    }

    /* Ẩn dòng dung lượng/định dạng mặc định */
    [data-testid="stFileUploaderDropzoneInstructions"] > div:last-child,
    [data-testid="stSidebar"] [data-testid="stFileUploader"] small {
        display:none !important;
        visibility:hidden !important;
        height:0 !important;
        margin:0 !important;
        padding:0 !important;
    }

    /* Giữ phần tử chứa nút rộng toàn ô */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] section > div {
        width:100% !important;
        display:flex !important;
        align-items:center !important;
        justify-content:center !important;
    }

    /* Nút tải file giống hình mẫu */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] button {
        position:relative !important;
        width:100% !important;
        min-height:58px !important;
        height:58px !important;
        margin:0 !important;
        padding:0 20px !important;
        border:1px solid #ff3b5a !important;
        border-radius:13px !important;
        background:linear-gradient(180deg,#ec0033 0%,#df002f 100%) !important;
        color:#ffffff !important;
        box-shadow:
            0 8px 18px rgba(230,0,45,.22),
            inset 0 1px 0 rgba(255,255,255,.22) !important;
        cursor:pointer !important;
        display:flex !important;
        align-items:center !important;
        justify-content:center !important;
        gap:12px !important;
        overflow:hidden !important;
    }

    /* Ẩn nội dung mặc định của Streamlit */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] button > * {
        display:none !important;
    }

    /* Icon upload nét trắng */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] button::before {
        content:"";
        width:27px !important;
        height:27px !important;
        flex:0 0 27px !important;
        display:block !important;
        background-repeat:no-repeat !important;
        background-position:center !important;
        background-size:contain !important;
        background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='28' height='28' viewBox='0 0 24 24' fill='none' stroke='white' stroke-width='2.2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M16 16l-4-4-4 4'/%3E%3Cpath d='M12 12v9'/%3E%3Cpath d='M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3'/%3E%3Cpath d='M16 16h5'/%3E%3C/svg%3E") !important;
        pointer-events:none !important;
    }

    [data-testid="stSidebar"] [data-testid="stFileUploader"] button::after {
        content:"Tải file mới";
        position:static !important;
        display:block !important;
        width:auto !important;
        height:auto !important;
        margin:0 !important;
        padding:0 !important;
        color:#ffffff !important;
        font-family:Inter, "Segoe UI", Arial, sans-serif !important;
        font-size:1rem !important;
        font-weight:800 !important;
        letter-spacing:.1px !important;
        line-height:1 !important;
        white-space:nowrap !important;
        pointer-events:none !important;
    }

    [data-testid="stSidebar"] [data-testid="stFileUploader"] button:hover {
        filter:brightness(.97) !important;
        transform:translateY(-1px) !important;
        box-shadow:
            0 10px 22px rgba(230,0,45,.27),
            inset 0 1px 0 rgba(255,255,255,.24) !important;
    }

    [data-testid="stSidebar"] [data-testid="stMetric"] { display:none !important; }
    .side-section-title { margin:10px 0 10px; font-size:1rem; font-weight:800; color:#121b2d; }
    .side-divider { height:1px; background:#edf0f5; margin:14px 0; }
    .file-info-row { display:flex; align-items:center; gap:11px; margin:12px 3px 10px; padding:3px 0; color:#192235; }
    .file-info-main { min-width:0; flex:1; }
    .file-name { font-size:.91rem; font-weight:700; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .file-meta { margin-top:2px; font-size:.75rem; color:#7d899d; }
    .file-icon { width:29px; height:34px; flex:0 0 auto; color:#0aa957; }
    .trash-icon { width:20px; height:20px; color:#26334a; flex:0 0 auto; }
    .upload-red { display:flex; align-items:center; justify-content:center; gap:8px; width:100%; box-sizing:border-box;
        padding:10px 12px; margin:10px 0 6px; color:white; font-weight:750; font-size:.9rem; border-radius:8px;
        background:linear-gradient(90deg,#e6002d,#ee173e); box-shadow:0 7px 17px rgba(230,0,45,.18); }
    .upload-hint { text-align:center; font-size:.75rem; color:#8a96a9; margin-bottom:2px; }
    .stat-card { display:flex; align-items:center; gap:12px; padding:13px; margin:0 0 11px; background:#fff;
        border:1px solid #e5e9f0; border-radius:13px; box-shadow:0 4px 12px rgba(20,32,51,.045); }
    .stat-icon { width:39px; height:39px; border-radius:50%; display:grid; place-items:center; color:#fff; flex:0 0 auto; }
    .stat-icon svg { width:20px; height:20px; stroke:currentColor; }
    .stat-icon.red { background:#ed1238; } .stat-icon.orange { background:#ff970f; }
    .stat-icon.green { background:#10a95b; } .stat-icon.blue { background:#188bea; }
    .stat-label { color:#718096; font-size:.76rem; line-height:1.25; }
    .stat-value { color:#111a2e; font-size:1.55rem; line-height:1.1; font-weight:850; margin-top:2px; }
    .update-row { display:flex; align-items:center; gap:10px; padding:1px 4px 0; }
    .update-icon { width:24px; height:24px; color:#8290a5; flex:0 0 auto; }
    .update-label { font-size:.76rem; color:#8490a3; }
    .update-value { margin-top:2px; font-size:.82rem; font-weight:800; color:#ed1238; }

    [data-testid="stChatMessage"] {
        max-width:1040px; background:#fff; border:1px solid var(--line); border-radius:17px;
        padding:.95rem 1.05rem; margin:.8rem auto; box-shadow:0 6px 20px rgba(20,32,51,.05);
    }
    [data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarUser"]) {
        /* Câu nhập lệnh hiển thị bên trái */
        width:fit-content;
        max-width:72%;
        margin-left:0 !important;
        margin-right:auto !important;
        background:linear-gradient(135deg,var(--brand-dark),var(--brand)) !important;
        border:0 !important;
        box-shadow:0 9px 24px rgba(230,0,45,.20);
    }
    [data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarUser"]) * { color:#fff !important; }
    [data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarAssistant"]) { border-left:3px solid var(--brand); }

    [data-testid="stBottom"] {
        background:linear-gradient(180deg,rgba(247,249,252,0),var(--page) 30%);
        padding-top:1.2rem;
        padding-bottom:.55rem;
    }

    /* Dòng gợi ý nằm ngay dưới ô nhập câu hỏi */
    [data-testid="stBottom"]::after {
        content:"💡 Gợi ý: Bạn có thể hỏi bằng tiếng Việt, ví dụ: “Top 10 FT tồn nhiều nhất”, “WO quá hạn”, “Thống kê theo trạng thái”, “WO đóng hôm nay của binhnt59”…";
        display:block;
        max-width:1040px;
        margin:7px auto 0;
        padding:0 10px;
        box-sizing:border-box;
        color:#7b8495;
        font-size:.84rem;
        line-height:1.45;
        text-align:left;
    }

    [data-testid="stChatInput"] { max-width:1040px; margin:0 auto; background:transparent !important; border:0 !important; box-shadow:none !important; padding:0 !important; }
    [data-testid="stChatInput"] > div {
        background:#fff !important; border:1px solid #d9dee7 !important;
        border-radius:18px !important; box-shadow:0 12px 32px rgba(20,32,51,.11) !important;
        padding:6px 8px 6px 15px !important; min-height:58px !important;
    }
    [data-testid="stChatInput"] > div:focus-within { border-color:var(--brand) !important; box-shadow:0 0 0 3px rgba(230,0,45,.08),0 12px 32px rgba(20,32,51,.11) !important; }
    [data-testid="stChatInput"] textarea { color:var(--text)!important; caret-color:var(--brand)!important; font-size:.98rem!important; padding:9px 2px!important; }
    [data-testid="stChatInput"] textarea::placeholder { color:#9aa4b4!important; opacity:1!important; }
    [data-testid="stChatInput"] button { width:42px!important; height:42px!important; min-width:42px!important; border:0!important; border-radius:50%!important; color:#fff!important; background:var(--brand)!important; }
    [data-testid="stChatInput"] button:hover { background:var(--brand-dark)!important; }
    [data-testid="stChatInput"] button:disabled { color:#a4acba!important; background:#edf0f4!important; }

    .stButton>button {
        border-radius:10px;
        border:1px solid #dfe3ea;
        color:var(--text);
        background:#fff;
        font-weight:650;
    }
    .stButton>button:hover {
        color:var(--brand-dark);
        border-color:#f2a3a6;
        background:#fff5f6;
    }

    /* Nút xuất Excel: nền đỏ, chữ trắng, icon tải xuống gọn và chuyên nghiệp */
    .stDownloadButton > button {
        min-height:46px !important;
        padding:0 20px !important;
        display:inline-flex !important;
        align-items:center !important;
        justify-content:center !important;
        gap:10px !important;
        border:0 !important;
        border-radius:12px !important;
        background:linear-gradient(135deg,#e6002d 0%,#f41643 100%) !important;
        color:#fff !important;
        font-size:.94rem !important;
        font-weight:800 !important;
        letter-spacing:.01em !important;
        box-shadow:0 9px 22px rgba(230,0,45,.24) !important;
        transition:transform .18s ease, box-shadow .18s ease, filter .18s ease !important;
    }
    .stDownloadButton > button::before {
        content:"⇩";
        width:25px;
        height:25px;
        display:inline-flex;
        align-items:center;
        justify-content:center;
        border:1.5px solid rgba(255,255,255,.78);
        border-radius:7px;
        color:#fff;
        font-size:17px;
        font-weight:900;
        line-height:1;
        box-sizing:border-box;
    }
    .stDownloadButton > button:hover {
        color:#fff !important;
        border:0 !important;
        filter:brightness(.96) !important;
        transform:translateY(-1px) !important;
        box-shadow:0 12px 27px rgba(230,0,45,.31) !important;
    }
    .stDownloadButton > button:active {
        transform:translateY(0) !important;
        box-shadow:0 6px 15px rgba(230,0,45,.22) !important;
    }
    .stDownloadButton > button p,
    .stDownloadButton > button span,
    .stDownloadButton > button div {
        color:#fff !important;
        font-weight:800 !important;
    }
    [data-testid="stDataFrame"] { background:#fff; border:1px solid var(--line); border-radius:14px; overflow:hidden; }
    [data-testid="stAlert"] { border-radius:12px; }

    @media (max-width:900px) {
        [data-testid="stSidebar"] {
            width:260px !important;
            min-width:260px !important;
            max-width:260px !important;
        }
        .sidebar-viettel-logo-wrap { width:100%; height:96px; min-height:96px; }
        .block-container { padding:.75rem .75rem 7rem !important; }
        .app-header { padding:9px 12px; gap:8px; }
        .app-header h1 { font-size:1rem; }
        .app-header p { font-size:.72rem; }
        .author-line { font-size:.67rem; margin-top:3px; }
        .viettel-logo { width:76px; max-width:22vw; }
        .app-logo { width:34px; height:34px; font-size:15px; }
        .welcome-card { margin:1.6rem auto; padding:1.4rem; }
        .sample-grid { grid-template-columns:1fr; }
        [data-testid="stChatMessage"]:has([data-testid="stChatMessageAvatarUser"]) { max-width:92%; }
    }
    




    .status-legend-list{
        margin-top:-8px;
        padding:0 2px;
        display:grid;
        grid-template-columns:1fr;
        gap:0;
    }
    .status-legend-row{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap:8px;
        min-height:25px;
        padding:3px 0;
        border-bottom:1px solid #f0f2f6;
    }
    .status-legend-name{
        display:flex;
        align-items:center;
        min-width:0;
        gap:7px;
    }
    .status-dot{
        width:9px;
        height:9px;
        border-radius:50%;
        flex:0 0 auto;
    }
    .status-name-text{
        font-size:.68rem;
        color:#4b5563;
        white-space:nowrap;
        overflow:hidden;
        text-overflow:ellipsis;
    }
    .status-legend-value{
        font-size:.68rem;
        font-weight:800;
        color:#111827;
        white-space:nowrap;
    }


    .status-side-legend{
        display:flex;
        flex-direction:column;
        justify-content:center;
        gap:4px;
        min-height:180px;
        padding:4px 0 2px;
    }
    .status-side-row{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap:8px;
        padding:5px 0;
        border-bottom:1px solid #f0f2f6;
    }
    .status-side-name{
        display:flex;
        align-items:center;
        min-width:0;
        gap:7px;
        flex:1;
    }
    .status-side-value{
        text-align:right;
        font-size:.72rem;
        font-weight:800;
        color:#111827;
        line-height:1.05;
        white-space:nowrap;
    }
    .status-side-value small{
        font-size:.62rem;
        color:#7b8495;
        font-weight:700;
    }


    
    
    
    
    

    /* ===== Dashboard quản lý WO ===== */
    .dash-topbar{
        display:flex; align-items:center; justify-content:space-between; gap:18px;
        padding:16px 20px; margin:0 0 14px;
        background:linear-gradient(90deg,#d8002d,#f0183e);
        border-radius:16px; color:white;
        box-shadow:0 12px 28px rgba(216,0,45,.22);
    }
    .dash-title-wrap{display:flex;align-items:center;gap:14px;min-width:0}
    .dash-menu-icon{font-size:1.45rem;font-weight:800}
    .dash-topbar h2{margin:0;font-size:1.35rem;line-height:1.2;color:white}
    .dash-topbar p{margin:4px 0 0;color:rgba(255,255,255,.82);font-size:.82rem}
    .dash-clock{
        padding:9px 13px;border:1px solid rgba(255,255,255,.35);
        border-radius:10px;background:rgba(255,255,255,.08);
        font-weight:700;font-size:.82rem;white-space:nowrap
    }
    .kpi-grid{
        display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin-bottom:14px
    }
    .kpi-card{
        display:flex;align-items:center;gap:12px;background:white;border:1px solid var(--line);
        border-radius:15px;padding:14px;box-shadow:0 7px 18px rgba(20,32,51,.06);min-width:0
    }
    .kpi-icon{
        width:46px;height:46px;border-radius:13px;display:grid;place-items:center;
        color:white;font-size:1.25rem;font-weight:800;flex:0 0 auto
    }
    .kpi-icon.red{background:linear-gradient(135deg,#d8002d,#f0183e)}
    .kpi-icon.green{background:linear-gradient(135deg,#168c45,#33b96a)}
    .kpi-icon.orange{background:linear-gradient(135deg,#ef7b00,#ff9e23)}
    .kpi-icon.blue{background:linear-gradient(135deg,#1f5ca8,#3479d4)}
    .kpi-icon.darkred{background:linear-gradient(135deg,#b40024,#e6002d)}
    .kpi-label{font-size:.90rem;font-weight:800;color:#232b38;text-transform:uppercase;letter-spacing:.02em}
    .kpi-value{font-size:1.65rem;font-weight:900;line-height:1.05;margin-top:4px}
    .kpi-note{font-size:.68rem;color:#7b8495;margin-top:5px;white-space:nowrap}
    .panel-card{
        background:#fff;border:1px solid var(--line);border-radius:15px;
        padding:12px 14px;box-shadow:0 7px 18px rgba(20,32,51,.05);height:100%
    }
    .panel-title{font-size:.92rem;font-weight:850;color:#202838;margin:0 0 8px}
    .dashboard-table-title{font-size:1rem;font-weight:850;margin:8px 0 8px;color:#202838}
    [data-testid="stTabs"] button p{font-weight:800!important}
    @media(max-width:1250px){.kpi-grid{grid-template-columns:repeat(3,minmax(0,1fr))}}
    @media(max-width:760px){
        .kpi-grid{grid-template-columns:1fr 1fr}
        .dash-topbar{align-items:flex-start;flex-direction:column}
        .dash-clock{white-space:normal}
    }


    /* ===== Dashboard giống mẫu quản lý WO ===== */
    .block-container{
        padding-top:.35rem !important;
        padding-left:.75rem !important;
        padding-right:.75rem !important;
    }
    .dash-topbar{
        border-radius:0 !important;
        margin:-.35rem -.75rem 6px !important;
        padding:0 !important;
        min-height:5px !important;
        height:5px !important;
        background:linear-gradient(90deg,#d8002d,#f0183e) !important;
        box-shadow:none !important;
    }
    .dash-topbar h2,
    .dash-topbar p,
    .dash-menu-icon,
    .dash-clock,
    .dash-title-wrap{
        display:none !important;
    }
    .kpi-grid{
        grid-template-columns:repeat(5,minmax(0,1fr)) !important;
        gap:10px !important;
        margin-bottom:10px !important;
    }
    .kpi-card{
        border-radius:12px !important;
        padding:12px 13px !important;
        min-height:84px;
    }
    .kpi-icon{
        width:44px !important;
        height:44px !important;
        border-radius:11px !important;
    }
    .kpi-value{font-size:2rem !important}
    .kpi-note{font-size:.64rem !important}
    .panel-title,.chart-card-title{
        margin:0 0 6px !important;
        font-size:.86rem !important;
        font-weight:850;
        color:#202838;
    }
    [data-testid="stContainer"]{
        border-color:#e6eaf0 !important;
        border-radius:12px !important;
        background:#fff !important;
        box-shadow:0 4px 12px rgba(20,32,51,.05) !important;
    }
    .dashboard-table-title{
        margin:4px 0 6px !important;
        font-size:.92rem !important;
    }
    .sidebar-nav{
        margin:0 -4px 14px;
        display:flex;
        flex-direction:column;
        gap:4px;
    }
    .sidebar-nav-item{
        display:flex;
        align-items:center;
        gap:10px;
        padding:9px 11px;
        border-radius:9px;
        font-size:.82rem;
        color:#263247;
        font-weight:650;
    }
    .sidebar-nav-item.active{
        background:#fff0f3;
        color:#e6002d;
        border-left:3px solid #e6002d;
    }
    .sidebar-nav-icon{
        width:20px;
        text-align:center;
        font-size:1rem;
    }
    .sidebar-group-title{
        margin:14px 0 8px;
        color:#e6002d;
        font-size:.73rem;
        font-weight:850;
        text-transform:uppercase;
    }
    @media(max-width:1250px){
        .kpi-grid{grid-template-columns:repeat(3,minmax(0,1fr)) !important}
    }


    /* Giữ biểu đồ trạng thái nằm gọn trong card */
    [data-testid="stVerticalBlockBorderWrapper"]:has(.chart-card-title) {
        overflow:hidden !important;
    }
    [data-testid="stVerticalBlockBorderWrapper"]:has(.chart-card-title)
    [data-testid="stVegaLiteChart"] {
        margin-top:-4px !important;
        margin-bottom:-4px !important;
    }



    /* ===== Nút mở trình chọn file: không bao giờ hiển thị preview ===== */
    [data-testid="stSidebar"] .st-key-upload_action_box {
        border:1px solid #dfe5ee !important;
        border-radius:13px !important;
        padding:14px 12px !important;
        background:#fbfcfe !important;
        margin-bottom:4px !important;
    }

    [data-testid="stSidebar"] .st-key-upload_action_box button {
        width:100% !important;
        min-height:58px !important;
        border:1px solid #ff3b5a !important;
        border-radius:13px !important;
        background:linear-gradient(180deg,#ec0033 0%,#df002f 100%) !important;
        color:#ffffff !important;
        font-size:1rem !important;
        font-weight:800 !important;
        box-shadow:0 8px 18px rgba(230,0,45,.22) !important;
    }

    [data-testid="stSidebar"] .st-key-upload_action_box button:hover {
        color:#ffffff !important;
        filter:brightness(.97) !important;
        transform:translateY(-1px) !important;
    }

    [data-testid="stSidebar"] .st-key-upload_action_box button p {
        color:#ffffff !important;
        font-weight:800 !important;
    }

    /* ===== Help clickable questions ===== */
    .help-question .stButton{
        margin:0 !important;
    }

    .help-question .stButton>button{
        width:100% !important;
        justify-content:flex-start !important;
        text-align:left !important;
        min-height:26px !important;
        padding:2px 8px !important;
        margin:0 !important;
        border:none !important;
        background:transparent !important;
        box-shadow:none !important;
        border-radius:6px !important;
    }

    .help-question .stButton>button p{
        margin:0 !important;
        line-height:1.08 !important;
        font-size:0.93rem !important;
    }

    .help-question .stButton>button:hover{
        background:#fff3f5 !important;
        color:#e6002d !important;
    }
</style>
    """,
    unsafe_allow_html=True,
)

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_FILE = DATA_DIR / "WoExport.xlsx"
PARQUET_FILE = DATA_DIR / "WoExport.parquet"
DUCKDB_FILE = DATA_DIR / "wo_bot.duckdb"

# Chỉ giới hạn số dòng hiển thị trên giao diện.
# Dữ liệu xuất Excel vẫn giữ đầy đủ.
MAX_DISPLAY_ROWS = 300
MAX_CHAT_HISTORY = 1

# Chế độ Turbo: chỉ dựng trang người dùng đang mở,
# không dựng đồng thời Dashboard và Chatbot.

COL_WORK_CODE = "Mã công việc"
COL_PARENT_CODE = "Mã công việc cha"
COL_WORK_TYPE = "Loại công việc"
COL_CONTENT = "Nội dung công việc"
COL_DESCRIPTION = "Mô tả"
COL_STATUS = "Trạng thái"
COL_EMPLOYEE = "Nhân viên thực hiện"
COL_CREATED = "Thời điểm tạo"
COL_DEADLINE = "Thời điểm yêu cầu kết thúc (dd/MM/yyyy HH:mm:ss)"
COL_REMAINING = "Thời gian còn lại (H)"
COL_STATION = "Mã trạm"
COL_SUBSCRIBER = "Thuê bao"
COL_PRIORITY = "Mức độ ưu tiên"
COL_FT_FINISHED = "Thời điểm FT hoàn thành"
COL_CD_CLOSED = "Thời điểm CD đóng"
COL_SYSTEM = "Hệ thống"
COL_COORD_GROUP = "Nhóm điều phối"

# Các mã hệ thống được loại khỏi toàn bộ Dashboard, chatbot và file xuất.
EXCLUDED_SYSTEMS = {"SPM", "SPM_VTNET", "SAP_KTTS"}
# Các loại công việc bắt đầu bằng các tiền tố này cũng được loại khỏi dữ liệu.
EXCLUDED_WORK_TYPE_PREFIXES = ("SAP_",)

SEARCH_STATUS = "_search_status"
SEARCH_EMPLOYEE = "_search_employee"
SEARCH_COORD_GROUP = "_search_coord_group"
SEARCH_SYSTEM = "_search_system"
SEARCH_STATION = "_search_station"
SEARCH_WORK_CODE = "_search_work_code"
SEARCH_SUBSCRIBER = "_search_subscriber"
SEARCH_CONTENT = "_search_content"
SEARCH_DESCRIPTION = "_search_description"
SEARCH_WORK_TYPE = "_search_work_type"

CLOSED_STATUS = "Đóng"

STATUS_LIST = [
    "Đã giao FT",
    "Đóng",
    "FT Đang thực hiện",
    "Chờ CD tiếp nhận",
    "FT hoàn thành",
    "CD đã tiếp nhận",
    "FT từ chối",
    "CD từ chối",
]

DISPLAY_COLUMNS = [
    COL_WORK_CODE,
    COL_WORK_TYPE,
    COL_CONTENT,
    COL_STATUS,
    COL_EMPLOYEE,
    COL_PRIORITY,
    COL_CREATED,
    COL_DEADLINE,
    COL_REMAINING,
    COL_FT_FINISHED,
    COL_SYSTEM,
    COL_COORD_GROUP,
    COL_STATION,
    COL_SUBSCRIBER,
]

# Các cột nguồn duy nhất cần giữ trong kho dữ liệu tối ưu.
CORE_SOURCE_COLUMNS = [
    COL_WORK_CODE,
    COL_PARENT_CODE,
    COL_WORK_TYPE,
    COL_CONTENT,
    COL_DESCRIPTION,
    COL_STATUS,
    COL_EMPLOYEE,
    COL_CREATED,
    COL_DEADLINE,
    COL_REMAINING,
    COL_STATION,
    COL_SUBSCRIBER,
    COL_PRIORITY,
    COL_FT_FINISHED,
    COL_CD_CLOSED,
    COL_SYSTEM,
    COL_COORD_GROUP,
]


@st.cache_resource(show_spinner=False)
def get_duckdb_connection():
    """Một kết nối DuckDB dùng lại cho toàn bộ phiên Streamlit."""
    connection = duckdb.connect(str(DUCKDB_FILE))
    connection.execute("PRAGMA threads=4")
    connection.execute("PRAGMA enable_object_cache=true")
    return connection


def parquet_sql_source() -> str:
    """
    Nguồn Parquet dùng cho DuckDB.

    Loại các mã hệ thống không sử dụng và các loại công việc bắt đầu
    bằng SAP_ ngay tại nguồn. Điều này vẫn có hiệu lực khi file Parquet
    được tạo từ phiên bản chương trình cũ.
    """
    escaped_path = str(PARQUET_FILE).replace("'", "''")
    excluded_values = ", ".join(
        f"'{value.replace(chr(39), chr(39) * 2)}'"
        for value in sorted(EXCLUDED_SYSTEMS)
    )
    work_type_conditions = " AND ".join(
        f"COALESCE(UPPER(TRIM(CAST(\"{COL_WORK_TYPE}\" AS VARCHAR))), '') "
        f"NOT LIKE '{prefix.replace(chr(39), chr(39) * 2)}%'"
        for prefix in EXCLUDED_WORK_TYPE_PREFIXES
    )
    return (
        f"(SELECT * FROM read_parquet('{escaped_path}') "
        f"WHERE COALESCE(UPPER(TRIM(CAST(\"{COL_SYSTEM}\" AS VARCHAR))), '') "
        f"NOT IN ({excluded_values}) "
        f"AND {work_type_conditions})"
    )


def duckdb_df(sql: str, params=None) -> pd.DataFrame:
    """Chạy SQL và chỉ trả về bảng kết quả nhỏ cần thiết."""
    return get_duckdb_connection().execute(sql, params or []).df()


def duckdb_scalar(sql: str, params=None, default=0):
    row = get_duckdb_connection().execute(sql, params or []).fetchone()
    return row[0] if row and row[0] is not None else default


@st.cache_data(show_spinner=False, ttl=300)
def get_distinct_values(column: str, parquet_mtime: float) -> list[str]:
    """Danh sách giá trị lọc lấy trực tiếp bằng DuckDB."""
    source = parquet_sql_source()
    query = f"""
        SELECT DISTINCT CAST("{column}" AS VARCHAR) AS value
        FROM {source}
        WHERE "{column}" IS NOT NULL
          AND TRIM(CAST("{column}" AS VARCHAR)) <> ''
        ORDER BY value
    """
    return duckdb_df(query)["value"].astype(str).tolist()


@st.cache_data(show_spinner=False, ttl=300)
def get_sidebar_summary(
    parquet_mtime: float,
    where_sql: str = "",
    params: tuple = (),
) -> dict:
    """Các KPI sidebar dùng chung điều kiện lọc với Dashboard."""
    source = parquet_sql_source()
    query = f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN LOWER(TRIM(CAST("{COL_STATUS}" AS VARCHAR))) = LOWER(?) THEN 1 ELSE 0 END) AS closed,
            SUM(CASE WHEN LOWER(TRIM(CAST("{COL_STATUS}" AS VARCHAR))) <> LOWER(?) THEN 1 ELSE 0 END) AS opened,
            SUM(CASE WHEN CAST("{COL_FT_FINISHED}" AS DATE) = CURRENT_DATE THEN 1 ELSE 0 END) AS today_finished,
            SUM(CASE WHEN CAST("{COL_FT_FINISHED}" AS DATE) = CURRENT_DATE - 1 THEN 1 ELSE 0 END) AS yesterday_finished
        FROM {source}
        {where_sql}
    """
    row = get_duckdb_connection().execute(
        query,
        [CLOSED_STATUS, CLOSED_STATUS] + list(params),
    ).fetchone()
    return {
        "total": int(row[0] or 0),
        "closed": int(row[1] or 0),
        "opened": int(row[2] or 0),
        "today_finished": int(row[3] or 0),
        "yesterday_finished": int(row[4] or 0),
    }


def build_sql_where(
    date_column=None,
    selected_dates=None,
    systems=None,
    coord_groups=None,
    employees=None,
    statuses=None,
):
    """Tạo WHERE và tham số cho Dashboard."""
    conditions = []
    params = []

    if date_column and selected_dates:
        if isinstance(selected_dates, (tuple, list)) and len(selected_dates) == 2:
            start_date, end_date = selected_dates
        else:
            start_date = end_date = selected_dates
        conditions.append(f'CAST("{date_column}" AS DATE) BETWEEN ? AND ?')
        params.extend([start_date, end_date])

    for column, values in [
        (COL_SYSTEM, systems),
        (COL_COORD_GROUP, coord_groups),
        (COL_EMPLOYEE, employees),
        (COL_STATUS, statuses),
    ]:
        if values:
            placeholders = ",".join(["?"] * len(values))
            conditions.append(
                f'CAST("{column}" AS VARCHAR) IN ({placeholders})'
            )
            params.extend(values)

    return (
        (" WHERE " + " AND ".join(conditions)) if conditions else "",
        params,
    )


# ============================================================
# HÀM CHUẨN HÓA
# ============================================================

def normalize_text(value) -> str:
    """Chuyển chuỗi về chữ thường, bỏ dấu tiếng Việt và khoảng trắng thừa."""
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(
        char for char in text
        if unicodedata.category(char) != "Mn"
    )
    text = text.replace("đ", "d")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_code(value) -> str:
    """Chuẩn hóa mã WO, mã trạm, thuê bao."""
    text = normalize_text(value)
    return re.sub(r"[^a-z0-9_-]", "", text)


def safe_contains(series: pd.Series, keyword: str) -> pd.Series:
    """
    Tìm từ khóa trên chuỗi đã chuẩn hóa.
    Nếu series là cột _search_* thì không xử lý lại từng dòng.
    """
    keyword_normalized = normalize_text(keyword)

    if str(series.name).startswith("_search_"):
        normalized_series = series.fillna("").astype(str)
    else:
        normalized_series = (
            series.fillna("")
            .astype(str)
            .map(normalize_text)
        )

    return normalized_series.str.contains(
        re.escape(keyword_normalized),
        regex=True,
        na=False,
    )


def exclude_unwanted_systems(df: pd.DataFrame) -> pd.DataFrame:
    """Loại các WO thuộc hệ thống không dùng và loại công việc bắt đầu bằng SAP_."""
    if df.empty:
        return df

    keep_mask = pd.Series(True, index=df.index)

    if COL_SYSTEM in df.columns:
        normalized_system = (
            df[COL_SYSTEM]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        keep_mask &= ~normalized_system.isin(EXCLUDED_SYSTEMS)

    if COL_WORK_TYPE in df.columns:
        normalized_work_type = (
            df[COL_WORK_TYPE]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        for prefix in EXCLUDED_WORK_TYPE_PREFIXES:
            keep_mask &= ~normalized_work_type.str.startswith(prefix, na=False)

    return df.loc[keep_mask].copy()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Chuẩn hóa dữ liệu một lần ngay khi đọc file."""
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]

    # Loại bỏ cột không dùng để giảm mạnh RAM và dung lượng Parquet.
    available_core_columns = [
        column for column in CORE_SOURCE_COLUMNS
        if column in df.columns
    ]
    df = df[available_core_columns].copy()

    # Loại các hệ thống không dùng và loại công việc bắt đầu bằng SAP_.
    df = exclude_unwanted_systems(df)

    date_columns = [
        COL_CREATED,
        COL_DEADLINE,
        COL_FT_FINISHED,
        COL_CD_CLOSED,
    ]

    for column in date_columns:
        if column in df.columns:
            df[column] = pd.to_datetime(
                df[column],
                dayfirst=True,
                errors="coerce",
            )

    if COL_REMAINING in df.columns:
        df[COL_REMAINING] = pd.to_numeric(
            df[COL_REMAINING],
            errors="coerce",
        )

    # Tạo sẵn các cột tìm kiếm để các câu hỏi sau không phải
    # normalize_text() lại trên hàng trăm nghìn dòng.
    # Chỉ tạo cột chuẩn hóa cho các trường lọc thường xuyên.
    # Không tạo bản sao chuẩn hóa của Nội dung/Mô tả vì các cột này rất lớn,
    # làm tăng mạnh RAM và thời gian khởi động với 150.000+ dòng.
    search_columns = {
        COL_STATUS: SEARCH_STATUS,
        COL_EMPLOYEE: SEARCH_EMPLOYEE,
        COL_COORD_GROUP: SEARCH_COORD_GROUP,
        COL_SYSTEM: SEARCH_SYSTEM,
        COL_STATION: SEARCH_STATION,
        COL_WORK_CODE: SEARCH_WORK_CODE,
        COL_SUBSCRIBER: SEARCH_SUBSCRIBER,
        COL_WORK_TYPE: SEARCH_WORK_TYPE,
    }

    for source_column, search_column in search_columns.items():
        if source_column in df.columns:
            df[search_column] = (
                df[source_column]
                .fillna("")
                .astype(str)
                .map(normalize_text)
            )

    # Các cột phân loại có số lượng giá trị lặp lại nhiều:
    # category giúp giảm đáng kể bộ nhớ.
    categorical_columns = [
        COL_STATUS,
        COL_EMPLOYEE,
        COL_COORD_GROUP,
        COL_SYSTEM,
        COL_PRIORITY,
        COL_WORK_TYPE,
        SEARCH_STATUS,
        SEARCH_EMPLOYEE,
        SEARCH_COORD_GROUP,
        SEARCH_SYSTEM,
    ]
    for column in categorical_columns:
        if column in df.columns:
            try:
                df[column] = df[column].astype("category")
            except Exception:
                pass

    return df


@st.cache_resource(show_spinner=False)
def load_default_excel(path: str, modified_time: float) -> pd.DataFrame:
    """Chỉ dùng khi chưa có Parquet."""
    return clean_dataframe(pd.read_excel(path, header=7))


@st.cache_resource(show_spinner=False)
def load_parquet(path: str, modified_time: float) -> pd.DataFrame:
    """Đọc Parquet và loại các hệ thống không sử dụng."""
    df = pd.read_parquet(path, engine="pyarrow", dtype_backend="pyarrow")
    return exclude_unwanted_systems(df)


def save_parquet(df: pd.DataFrame) -> bool:
    """
    Lưu dữ liệu tối ưu. Trả về False nếu máy chưa cài pyarrow/fastparquet.
    """
    try:
        df.to_parquet(
            PARQUET_FILE,
            index=False,
            engine="pyarrow",
            compression="snappy",
        )
        return True
    except (ImportError, ModuleNotFoundError):
        return False


# ============================================================
# ĐỌC DỮ LIỆU
# ============================================================

def get_dataframe() -> pd.DataFrame:
    """
    Luồng dữ liệu tối ưu:
    - Upload Excel chỉ đọc một lần.
    - Sau đó tạo data/WoExport.parquet.
    - Các lần F5 và câu hỏi tiếp theo đọc Parquet.
    """
    if "show_excel_uploader" not in st.session_state:
        st.session_state["show_excel_uploader"] = False

    if "excel_uploader_version" not in st.session_state:
        st.session_state["excel_uploader_version"] = 0

    if not st.session_state["show_excel_uploader"]:
        with st.sidebar.container(key="upload_action_box"):
            open_uploader = st.button(
                "☁  Tải file mới",
                key="open_excel_uploader",
                use_container_width=True,
                type="primary",
            )

        if open_uploader:
            st.session_state["show_excel_uploader"] = True
            st.rerun()

    uploaded_file = None

    if st.session_state["show_excel_uploader"]:
        uploaded_file = st.sidebar.file_uploader(
            "Chọn file Excel mới",
            type=["xlsx"],
            label_visibility="collapsed",
            key=f"wo_excel_uploader_{st.session_state['excel_uploader_version']}",
        )

        cancel_upload = st.sidebar.button(
            "Hủy chọn file",
            key="cancel_excel_uploader",
            use_container_width=True,
        )

        if cancel_upload:
            st.session_state["show_excel_uploader"] = False
            st.session_state["excel_uploader_version"] += 1
            st.rerun()

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()

        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)

            temp_file = DATA_DIR / "WoExport.uploading.xlsx"
            temp_file.write_bytes(file_bytes)
            temp_file.replace(DEFAULT_FILE)

            with st.spinner("Đang tối ưu dữ liệu lần đầu..."):
                uploaded_df = clean_dataframe(
                    pd.read_excel(io.BytesIO(file_bytes), header=7)
                )

                parquet_created = save_parquet(uploaded_df)

        except Exception as error:
            st.sidebar.error(
                "Không thể cập nhật dữ liệu.\n\n"
                f"Đường dẫn: {DEFAULT_FILE}\n\n"
                f"Lỗi: {error}"
            )
            st.stop()

        try:
            load_default_excel.clear()
            load_parquet.clear()
            get_distinct_values.clear()
            get_sidebar_summary.clear()
        except Exception:
            pass
        st.session_state["show_excel_uploader"] = False
        st.session_state["excel_uploader_version"] += 1
        st.session_state["upload_success"] = True
        st.session_state["parquet_created"] = parquet_created
        st.rerun()

    if st.session_state.pop("upload_success", False):
        if st.session_state.pop("parquet_created", False):
            st.sidebar.success(
                "Dữ liệu đã cập nhật và tối ưu tốc độ.",
                icon="✅",
            )
        else:
            st.sidebar.warning(
                "Đã cập nhật Excel nhưng chưa tạo được Parquet. "
                "Hãy chạy: pip install pyarrow",
                icon="⚠️",
            )

    # Ưu tiên đọc Parquet vì nhanh hơn XLSX.
    if PARQUET_FILE.exists():
        return load_parquet(
            str(PARQUET_FILE),
            PARQUET_FILE.stat().st_mtime,
        )

    if DEFAULT_FILE.exists():
        with st.spinner("Đang tạo dữ liệu tối ưu lần đầu..."):
            df = load_default_excel(
                str(DEFAULT_FILE),
                DEFAULT_FILE.stat().st_mtime,
            )

            if save_parquet(df):
                try:
                    load_default_excel.clear()
                    load_parquet.clear()
                    get_distinct_values.clear()
                    get_sidebar_summary.clear()
                except Exception:
                    pass
                return load_parquet(
                    str(PARQUET_FILE),
                    PARQUET_FILE.stat().st_mtime,
                )

            st.sidebar.warning(
                "Máy chưa có pyarrow nên bot vẫn đang đọc Excel. "
                "Cài bằng: pip install pyarrow",
                icon="⚠️",
            )
            return df

    st.error(
        "Không tìm thấy file dữ liệu.\n\n"
        f"Hãy bấm **Tải file mới** hoặc chép file vào:\n{DEFAULT_FILE}"
    )
    st.stop()


# ============================================================
# NHẬN DIỆN Ý ĐỊNH
# ============================================================

def detect_intent(question: str) -> str:
    q = normalize_text(question)

    # Khi người dùng gửi "help" hoặc các cụm tương đương,
    # chatbot hiển thị hướng dẫn và câu hỏi mẫu.
    if q in {
        "help",
        "tro giup",
        "huong dan",
        "goi y",
        "cau hoi mau",
    }:
        return "help"

    if any(word in q for word in [
        "ve bieu do",
        "bieu do dong wo",
        "dashboard",
    ]):
        return "dashboard"

    if any(word in q for word in [
        "top",
        "nhieu nhat",
        "cao nhat",
        "ton nhieu nhat",
        "xep hang",
    ]):
        return "top"

    if any(word in q for word in [
        "thong ke",
        "theo trang thai",
        "theo tung trang thai",
        "tung trang thai",
        "phan bo",
        "tong hop",
        "bao cao",
        "tung nhan vien",
        "moi nhan vien",
        "theo nhan vien",
        "theo tung nhan vien",
        "theo ma he thong",
        "theo he thong",
        "nhom dieu phoi",
        "theo nhom dieu phoi",
        "tung nhom dieu phoi",
        "moi nhom dieu phoi",
    ]):
        return "statistics"

    if any(word in q for word in [
        "bao nhieu",
        "so luong",
        "dem",
        "tong wo",
        "co may",
    ]):
        return "count"

    if any(word in q for word in [
        "liet ke",
        "danh sach",
        "chi tiet",
        "thong tin",
        "tra cuu",
        "tim",
        "wo cua",
        "wo tram",
        "wo thue bao",
    ]):
        return "list"

    return "list"


def extract_top_number(question: str, default: int = 10) -> int:
    q = normalize_text(question)

    patterns = [
        r"\btop\s*(\d+)\b",
        r"\b(\d+)\s*(?:ft|nhan vien)\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, q)
        if match:
            return max(1, min(int(match.group(1)), 100))

    return default


# ============================================================
# TRÍCH XUẤT ĐIỀU KIỆN
# ============================================================

def extract_status(question: str):
    """Trích xuất trạng thái WO từ câu hỏi tiếng Việt."""
    q = normalize_text(question)

    status_aliases = [
        ("Đóng", [
            "wo dong",
            "da dong",
            "trang thai dong",
        ]),
        ("FT hoàn thành", [
            "ft da hoan thanh",
            "ft hoan thanh",
        ]),
        ("FT Đang thực hiện", [
            "ft dang thuc hien",
            "dang thuc hien",
            "dang xu ly",
        ]),
        ("Chờ CD tiếp nhận", [
            "cho cd tiep nhan",
            "cho dieu phoi tiep nhan",
        ]),
        ("CD đã tiếp nhận", [
            "cd da tiep nhan",
            "dieu phoi da tiep nhan",
        ]),
        ("FT từ chối", [
            "ft tu choi",
        ]),
        ("CD từ chối", [
            "cd tu choi",
            "dieu phoi tu choi",
        ]),
        ("Đã giao FT", [
            "da giao ft",
            "giao ft",
            "chua tiep nhan",
        ]),
    ]

    for actual_status, aliases in status_aliases:
        if any(alias in q for alias in aliases):
            return actual_status

    return None


def extract_station(question: str, df: pd.DataFrame):
    if COL_STATION not in df.columns:
        return None

    q = normalize_text(question)

    patterns = [
        r"(?:ma\s*)?tram\s*[:\-]?\s*([a-z0-9_-]+)",
        r"\b([a-z]{2,6}\d{2,})\b",
    ]

    for pattern in patterns:
        matches = re.findall(pattern, q)

        for candidate in matches:
            candidate_code = normalize_code(candidate)

            if candidate_code in {
                "wo", "ft", "cd", "top", "tram", "ton"
            }:
                continue

            station_values = (
                df[COL_STATION]
                .dropna()
                .astype(str)
                .map(normalize_code)
            )

            if station_values.str.contains(
                re.escape(candidate_code),
                regex=True,
            ).any():
                return candidate

    return None


def extract_work_code(question: str, df: pd.DataFrame):
    if COL_WORK_CODE not in df.columns:
        return None

    original_question = str(question)

    patterns = [
        r"\b(WO[_\-][A-Za-z0-9_\-]+)\b",
        r"\b([A-Za-z]+[_\-]WO[_\-][A-Za-z0-9_\-]+)\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, original_question, re.IGNORECASE)
        if match:
            return match.group(1)

    q = normalize_text(question)

    if any(phrase in q for phrase in [
        "ma wo",
        "ma cong viec",
        "tra wo",
        "thong tin wo",
    ]):
        match = re.search(
            r"(?:ma wo|ma cong viec|tra wo|thong tin wo)"
            r"\s*[:\-]?\s*([a-z0-9_-]+)",
            q,
        )

        if match:
            candidate = match.group(1)

            blocked_words = {
                "ton",
                "tram",
                "qua",
                "han",
                "hom",
                "nay",
                "nhan",
                "vien",
            }

            if candidate not in blocked_words:
                return candidate

    return None


def extract_subscriber(question: str):
    q = normalize_text(question)

    match = re.search(
        r"(?:thue bao|tb)\s*[:\-]?\s*([a-z0-9_-]+)",
        q,
    )

    if match:
        return match.group(1)

    return None


@st.cache_data(show_spinner=False)
def employee_candidates_from_values(values: tuple) -> list[str]:
    return [str(value).strip() for value in values if str(value).strip()]


def employee_candidates(df: pd.DataFrame):
    if COL_EMPLOYEE not in df.columns:
        return []
    values = tuple(df[COL_EMPLOYEE].dropna().astype(str).unique().tolist())
    return employee_candidates_from_values(values)


def extract_employee(question: str, df: pd.DataFrame):
    q = normalize_text(question)
    employees = employee_candidates(df)

    if not employees:
        return None

    # Ưu tiên khớp nguyên tên hoặc tài khoản xuất hiện trong câu hỏi
    matches = []

    for employee in employees:
        normalized_employee = normalize_text(employee)

        if normalized_employee and normalized_employee in q:
            matches.append((len(normalized_employee), employee))

    if matches:
        matches.sort(reverse=True)
        return matches[0][1]

    # Tìm sau các cụm "của", "nhân viên", "ft"
    patterns = [
        r"(?:cua|nhan vien|ft)\s+([a-z0-9_.\-]+)",
        r"([a-z0-9_.\-]+)\s+con\s+bao nhieu",
    ]

    for pattern in patterns:
        match = re.search(pattern, q)

        if not match:
            continue

        candidate = match.group(1)

        blocked_words = {
            "tram",
            "thue",
            "bao",
            "hom",
            "nay",
            "qua",
            "han",
            "dang",
            "hoan",
            "thanh",
        }

        if candidate in blocked_words:
            continue

        possible_matches = []

        for employee in employees:
            employee_normalized = normalize_text(employee)

            if (
                candidate in employee_normalized
                or employee_normalized in candidate
            ):
                possible_matches.append(employee)

        if len(possible_matches) == 1:
            return possible_matches[0]

    return None



@st.cache_data(show_spinner=False)
def coordination_group_candidates_from_values(values: tuple) -> list[str]:
    return [str(value).strip() for value in values if str(value).strip()]


def coordination_group_candidates(df: pd.DataFrame):
    """Danh sách Nhóm điều phối có trong dữ liệu."""
    if COL_COORD_GROUP not in df.columns:
        return []
    values = tuple(df[COL_COORD_GROUP].dropna().astype(str).unique().tolist())
    return coordination_group_candidates_from_values(values)


def extract_coordination_group(question: str, df: pd.DataFrame):
    """Nhận diện Nhóm điều phối được nhắc trong câu hỏi."""
    q = normalize_text(question)
    groups = coordination_group_candidates(df)

    if not groups:
        return None

    # Ưu tiên khớp trực tiếp tên đầy đủ hoặc mã trong ngoặc.
    matches = []
    for group in groups:
        normalized_group = normalize_text(group)

        if normalized_group and normalized_group in q:
            matches.append((len(normalized_group), group))
            continue

        # Cho phép hỏi bằng mã như VCC_GLI_005_IGI.
        codes = re.findall(r"\(([A-Za-z0-9_-]+)\)", group)
        for code in codes:
            if normalize_text(code) in q:
                matches.append((len(code), group))

    if matches:
        matches.sort(reverse=True)
        return matches[0][1]

    # Tìm phần đứng sau cụm "nhóm điều phối".
    match = re.search(
        r"(?:nhom dieu phoi|nhom dp|dieu phoi)"
        r"\s*[:\-]?\s*(.+?)(?:\s+(?:hom nay|hom qua|qua han|"
        r"theo trang thai|theo nhan vien|theo he thong|trong \d+ ngay)|$)",
        q,
    )

    if not match:
        return None

    candidate = match.group(1).strip()
    possible = []

    for group in groups:
        normalized_group = normalize_text(group)
        if candidate and (
            candidate in normalized_group
            or normalized_group in candidate
        ):
            possible.append(group)

    if len(possible) == 1:
        return possible[0]

    return None

@st.cache_data(show_spinner=False)
def work_type_candidates_from_values(values: tuple) -> list[str]:
    return [str(value).strip() for value in values if str(value).strip()]


def work_type_candidates(df: pd.DataFrame):
    """Danh sách loại công việc có trong dữ liệu."""
    if COL_WORK_TYPE not in df.columns:
        return []
    values = tuple(df[COL_WORK_TYPE].dropna().astype(str).unique().tolist())
    return work_type_candidates_from_values(values)


def extract_work_type(question: str, df: pd.DataFrame):
    """
    Nhận diện Loại công việc theo nguyên tắc an toàn:

    1. Khi câu hỏi có cụm chỉ rõ ``loại công việc``/``loại WO``, lấy phần
       người dùng nhập và lọc theo phép chứa. Vì vậy ``ICMS_Kiểm tra`` sẽ
       khớp tất cả ICMS_Kiểm tra nhà trạm, tuyến cáp, THC...
    2. Khi không có cụm chỉ rõ, chỉ tự nhận diện nếu câu hỏi có tiền tố đặc
       trưng thật sự xuất hiện ở đầu tên loại công việc, ví dụ ICMS, VTNET.
       Không dùng các từ nghiệp vụ chung như ``thực hiện``, ``trạng thái``,
       ``theo``, ``hệ thống`` làm điều kiện Loại công việc.
    """
    if COL_WORK_TYPE not in df.columns:
        return None

    q = normalize_text(question)
    values = work_type_candidates(df)
    if not values:
        return None

    normalized_values = [normalize_text(value) for value in values]
    compact_values = [re.sub(r"[^a-z0-9]+", "", value) for value in normalized_values]

    # Trường hợp người dùng chỉ rõ trường cần lọc.
    marker_match = re.search(
        r"(?:loai cong viec|loai wo|cong viec loai)"
        r"\s*[:\-]?\s*(.+?)"
        r"(?=\s+(?:duoc tao|tao trong|trong thang|thang \d+|nam \d{4}|"
        r"trang thai|dang thuc hien|da dong|ft hoan thanh|cua ft|cua nhan vien|"
        r"he thong|nhom dieu phoi|ma tram|tram|thue bao|qua han)\b|$)",
        q,
    )
    if marker_match:
        candidate = marker_match.group(1).strip(" _-")
        candidate_compact = re.sub(r"[^a-z0-9]+", "", candidate)
        if candidate_compact and any(
            candidate_compact in compact_value
            for compact_value in compact_values
        ):
            return candidate
        return None

    # Không có marker: tuyệt đối không suy đoán chỉ từ các từ nghiệp vụ chung.
    # Chỉ nhận khi có tiền tố đặc trưng ở đầu tên loại công việc, ví dụ ICMS_.
    blocked_prefixes = {
        "wo", "ft", "cd", "he", "trang", "nhom", "ma", "loai",
        "cong", "viec", "danh", "thong", "bao", "tao", "kiem",
    }
    prefixes = set()
    for value in normalized_values:
        prefix_match = re.match(r"\s*([a-z0-9]{3,})[_\-]", value)
        if prefix_match:
            prefix = prefix_match.group(1)
            if prefix not in blocked_prefixes:
                prefixes.add(prefix)

    found_prefix = None
    found_position = None
    for prefix in sorted(prefixes, key=len, reverse=True):
        match = re.search(rf"(?<![a-z0-9]){re.escape(prefix)}(?![a-z0-9])", q)
        if match:
            found_prefix = prefix
            found_position = match.start()
            break

    if not found_prefix:
        return None

    # Lấy cụm từ tiền tố đến trước điều kiện khác trong câu hỏi.
    candidate_text = q[found_position:]
    candidate_text = re.split(
        r"\b(?:duoc tao|tao trong|trong thang|thang \d+|nam \d{4}|"
        r"trang thai|dang thuc hien|da dong|ft hoan thanh|cua ft|cua nhan vien|"
        r"he thong|nhom dieu phoi|ma tram|tram|thue bao|qua han)\b",
        candidate_text,
        maxsplit=1,
    )[0].strip(" _-")

    # Loại các từ hỏi ở cuối nếu người dùng đặt câu ngắn.
    candidate_text = re.sub(
        r"\b(?:bao nhieu|so luong|danh sach|liet ke|thong ke)\b",
        " ",
        candidate_text,
    )
    candidate_text = re.sub(r"\s+", " ", candidate_text).strip(" _-")
    candidate_compact = re.sub(r"[^a-z0-9]+", "", candidate_text)

    if candidate_compact and any(
        candidate_compact in compact_value
        for compact_value in compact_values
    ):
        return candidate_text

    # Ít nhất vẫn cho phép lọc theo riêng tiền tố đặc trưng, ví dụ "WO ICMS".
    if any(
        re.search(rf"(?:^|[^a-z0-9]){re.escape(found_prefix)}[_\-]", value)
        for value in normalized_values
    ):
        return found_prefix

    return None

def extract_month_year(question: str):
    """Lấy tháng/năm trong câu hỏi; nếu thiếu năm thì dùng năm hiện tại."""
    q = normalize_text(question)

    patterns = [
        r"\bthang\s*(\d{1,2})(?:\s*(?:nam|/|-)\s*(\d{4}))?\b",
        r"\b(\d{1,2})[/-](\d{4})\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, q)
        if not match:
            continue

        month = int(match.group(1))
        year = int(match.group(2)) if match.group(2) else date.today().year
        if 1 <= month <= 12:
            return month, year

    return None


def extract_remaining_hours(question: str):
    q = normalize_text(question)

    patterns = [
        r"(?:con|duoi|nho hon)\s*(?:duoi)?\s*(\d+(?:[.,]\d+)?)\s*gio",
        r"trong\s*(\d+(?:[.,]\d+)?)\s*gio",
    ]

    for pattern in patterns:
        match = re.search(pattern, q)

        if match:
            return float(match.group(1).replace(",", "."))

    return None


def extract_specific_date(question: str):
    match = re.search(
        r"\b(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})\b",
        question,
    )

    if not match:
        return None

    day, month, year = map(int, match.groups())

    try:
        return date(year, month, day)
    except ValueError:
        return None


def extract_recent_days(question: str):
    """Lấy số ngày gần nhất, ví dụ: trong 3 ngày, 5 ngày qua."""
    q = normalize_text(question)

    patterns = [
        r"(?:trong|gan nhat|vong)\s*(\d+)\s*ngay",
        r"(\d+)\s*ngay\s*(?:gan nhat|qua|tro lai day)",
    ]

    for pattern in patterns:
        match = re.search(pattern, q)
        if match:
            days = int(match.group(1))
            return max(1, min(days, 365))

    return None


def extract_content_keyword(question: str):
    q = normalize_text(question)

    content_aliases = [
        "mat dien",
        "nguon",
        "truyen dan",
        "bao duong",
        "bts",
        "may no",
        "ac quy",
        "dieu hoa",
        "canh bao",
        "mang",
        "cap quang",
        "suy hao",
    ]

    # Các nội dung có sẵn
    for alias in content_aliases:
        if alias in q:
            return alias

    # Tìm phần nội dung đứng sau các cụm này
    match = re.search(
        r"(?:co\s+noi dung|noi dung|mo ta|lien quan|chua tu)"
        r"\s*[:\-]?\s*[\"']?(.+?)[\"']?$",
        q,
    )

    if not match:
        return None

    keyword = match.group(1).strip()

    # Cắt bỏ các điều kiện xuất hiện phía sau từ khóa nội dung
    split_patterns = [
        r"\s+cua\s+nhan vien\s+thuc hien\b",
        r"\s+cua\s+nhan vien\b",
        r"\s+cua\s+ft\b",
        r"\s+theo\s+nhan vien\b",
        r"\s+nhan vien\s+thuc hien\b",
        r"\s+nhan vien\b",
        r"\s+ma\s+tram\b",
        r"\s+tram\b",
        r"\s+thue\s+bao\b",
        r"\s+trang\s+thai\b",
        r"\s+qua\s+han\b",
        r"\s+hom\s+nay\b",
        r"\s+hom\s+qua\b",
        r"\s+thang\s+nay\b",
        r"\s+con\s+duoi\b",
        r"\s+con\s+toi\s+da\b",
    ]

    for pattern in split_patterns:
        keyword = re.split(
            pattern,
            keyword,
            maxsplit=1,
        )[0].strip()

    # Loại bỏ dấu nháy và khoảng trắng thừa
    keyword = keyword.strip("\"' ")
    keyword = re.sub(r"\s+", " ", keyword).strip()

    if len(keyword) >= 2:
        return keyword

    return None


# ============================================================
# ÁP DỤNG BỘ LỌC
# ============================================================

def apply_filters(df: pd.DataFrame, question: str):
    result = df
    filters = []
    q = normalize_text(question)

    # WO tồn
    asks_open = any(phrase in q for phrase in [
        "wo ton",
        "ton wo",
        "con bao nhieu wo",
        "chua dong",
        "dang ton",
        "tong wo ton",
    ])

    if asks_open and COL_STATUS in result.columns:
        status_series = (
            result[SEARCH_STATUS]
            if SEARCH_STATUS in result.columns
            else result[COL_STATUS].fillna("").astype(str).map(normalize_text)
        )
        result = result[status_series != normalize_text(CLOSED_STATUS)]
        filters.append("WO tồn")

    # Trạng thái
    status = extract_status(question)

    if status and COL_STATUS in result.columns:
        status_series = (
            result[SEARCH_STATUS]
            if SEARCH_STATUS in result.columns
            else result[COL_STATUS].fillna("").astype(str).map(normalize_text)
        )

        result = result[status_series == normalize_text(status)]

        filters.append(f"Trạng thái: {status}")

    # Quá hạn
    if any(phrase in q for phrase in [
        "qua han",
        "am gio",
        "tre han",
    ]) and COL_REMAINING in result.columns:
        result = result[result[COL_REMAINING] < 0]
        filters.append("Quá hạn")

    # Còn dưới N giờ
    remaining_hours = extract_remaining_hours(question)
    if remaining_hours is not None and COL_REMAINING in result.columns:
        result = result[
            result[COL_REMAINING].notna()
            & (result[COL_REMAINING] >= 0)
            & (result[COL_REMAINING] <= remaining_hours)
        ]
        filters.append(f"Còn tối đa {remaining_hours:g} giờ")

    # Nhân viên
    employee = extract_employee(question, df)
    if employee and COL_EMPLOYEE in result.columns:
        employee_series = (
            result[SEARCH_EMPLOYEE]
            if SEARCH_EMPLOYEE in result.columns
            else result[COL_EMPLOYEE].fillna("").astype(str).map(normalize_text)
        )
        result = result[employee_series == normalize_text(employee)]
        filters.append(f"Nhân viên: {employee}")

    # Nhóm điều phối
    coordination_group = extract_coordination_group(question, df)
    if coordination_group and COL_COORD_GROUP in result.columns:
        group_series = (
            result[SEARCH_COORD_GROUP]
            if SEARCH_COORD_GROUP in result.columns
            else result[COL_COORD_GROUP].fillna("").astype(str).map(normalize_text)
        )
        result = result[
            group_series == normalize_text(coordination_group)
        ]
        filters.append(f"Nhóm điều phối: {coordination_group}")

    # Hệ thống: người dùng nhập trực tiếp mã trong câu hỏi, ví dụ CC_SCVT.
    system_name = extract_system(question, df)
    if system_name and COL_SYSTEM in result.columns:
        system_series = (
            result[SEARCH_SYSTEM]
            if SEARCH_SYSTEM in result.columns
            else result[COL_SYSTEM].fillna("").astype(str).map(normalize_text)
        )
        result = result[
            system_series == normalize_text(system_name)
        ]
        filters.append(f"Hệ thống: {system_name}")

    # Loại công việc
    work_type = extract_work_type(question, df)
    if work_type and COL_WORK_TYPE in result.columns:
        work_type_series = (
            result[SEARCH_WORK_TYPE]
            if SEARCH_WORK_TYPE in result.columns
            else result[COL_WORK_TYPE].fillna("").astype(str).map(normalize_text)
        )
        # Tìm theo cụm để một câu hỏi ngắn như "ICMS_Kiểm tra" lấy đủ:
        # ICMS_Kiểm tra nhà trạm, ICMS_Kiểm tra tuyến cáp, ICMS_Kiểm tra THC...
        result = result[safe_contains(work_type_series, work_type)]
        filters.append(f"Loại công việc chứa: {work_type}")

    # Mã trạm
    station = extract_station(question, df)
    if station and COL_STATION in result.columns:
        search_series = (
            result[SEARCH_STATION]
            if SEARCH_STATION in result.columns
            else result[COL_STATION]
        )
        result = result[safe_contains(search_series, station)]
        filters.append(f"Mã trạm: {station}")

    # Mã WO
    work_code = extract_work_code(question, df)
    if work_code and COL_WORK_CODE in result.columns:
        search_series = (
            result[SEARCH_WORK_CODE]
            if SEARCH_WORK_CODE in result.columns
            else result[COL_WORK_CODE]
        )
        result = result[safe_contains(search_series, work_code)]
        filters.append(f"Mã WO: {work_code}")

    # Thuê bao
    subscriber = extract_subscriber(question)
    if subscriber and COL_SUBSCRIBER in result.columns:
        search_series = (
            result[SEARCH_SUBSCRIBER]
            if SEARCH_SUBSCRIBER in result.columns
            else result[COL_SUBSCRIBER]
        )
        result = result[safe_contains(search_series, subscriber)]
        filters.append(f"Thuê bao: {subscriber}")

    # Lọc theo ngày
    specific_date = extract_specific_date(question)
    recent_days = extract_recent_days(question)
    month_year = extract_month_year(question)

    asks_closed_date = any(phrase in q for phrase in [
        "wo dong",
        "da dong",
        "dong ngay",
        "dong hom nay",
        "dong hom qua",
        "thoi diem dong",
    ])
    asks_ft_finished_date = any(phrase in q for phrase in [
        "ft hoan thanh ngay",
        "ft hoan thanh hom nay",
        "ft hoan thanh hom qua",
        "hoan thanh ngay",
        "hoan thanh hom nay",
        "thoi diem ft hoan thanh",
    ])

    asks_created_date = any(phrase in q for phrase in [
        "duoc tao",
        "tao trong",
        "tao thang",
        "thoi diem tao",
        "ngay tao",
    ])

    # Khi người dùng nói rõ "được tạo" thì luôn lọc theo Thời điểm tạo.
    # Các câu hỏi về WO đóng/FT hoàn thành vẫn dùng Thời điểm FT hoàn thành.
    if asks_created_date:
        date_column = COL_CREATED
        date_label = "Thời điểm tạo"
    elif status == "Đóng" or asks_closed_date or asks_ft_finished_date:
        date_column = COL_FT_FINISHED
        date_label = "Thời điểm FT hoàn thành"
    else:
        date_column = COL_CREATED
        date_label = "Thời điểm tạo"

    if specific_date and date_column in result.columns:
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.date == specific_date)
        ]
        filters.append(
            f"{date_label} ngày {specific_date.strftime('%d/%m/%Y')}"
        )
    elif "hom nay" in q and date_column in result.columns:
        today = date.today()
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.date == today)
        ]
        filters.append(f"{date_label} hôm nay")
    elif "hom qua" in q and date_column in result.columns:
        yesterday = date.today() - timedelta(days=1)
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.date == yesterday)
        ]
        filters.append(f"{date_label} hôm qua")
    elif recent_days is not None and date_column in result.columns:
        # Tính cả ngày hôm nay. Ví dụ 3 ngày = hôm nay và 2 ngày trước.
        end_date = date.today()
        start_date = end_date - timedelta(days=recent_days - 1)
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.date >= start_date)
            & (result[date_column].dt.date <= end_date)
        ]
        filters.append(
            f"{date_label} trong {recent_days} ngày "
            f"({start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')})"
        )
    elif month_year is not None and date_column in result.columns:
        month, year = month_year
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.year == year)
            & (result[date_column].dt.month == month)
        ]
        filters.append(f"{date_label} trong tháng {month}/{year}")
    elif "thang nay" in q and date_column in result.columns:
        today = date.today()
        result = result[
            result[date_column].notna()
            & (result[date_column].dt.year == today.year)
            & (result[date_column].dt.month == today.month)
        ]
        filters.append(f"{date_label} trong tháng này")

    # Nội dung
    content_keyword = extract_content_keyword(question)
    if content_keyword:
        content_mask = pd.Series(False, index=result.index)

        content_search_columns = [
            (COL_CONTENT, SEARCH_CONTENT),
            (COL_DESCRIPTION, SEARCH_DESCRIPTION),
            (COL_WORK_TYPE, SEARCH_WORK_TYPE),
        ]

        for source_column, search_column in content_search_columns:
            if search_column in result.columns:
                content_mask = content_mask | safe_contains(
                    result[search_column],
                    content_keyword,
                )
            elif source_column in result.columns:
                content_mask = content_mask | safe_contains(
                    result[source_column],
                    content_keyword,
                )

        result = result[content_mask]
        filters.append(f"Nội dung: {content_keyword}")

    return result, filters


# ============================================================
# TRẢ LỜI
# ============================================================

def visible_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Loại các cột kỹ thuật _search_* khỏi dữ liệu hiển thị/xuất."""
    return df[
        [column for column in df.columns if not str(column).startswith("_search_")]
    ].copy()


def dataframe_to_excel(df: pd.DataFrame) -> bytes:
    """Xuất Excel có sẵn định dạng dễ đọc, không thay đổi dữ liệu."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    export_df = visible_dataframe(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            index=False,
            sheet_name="KetQua",
        )

        worksheet = writer.sheets["KetQua"]
        max_row = worksheet.max_row
        max_column = worksheet.max_column

        # Cố định hàng tiêu đề và bật bộ lọc.
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        # Định dạng tiêu đề theo màu thương hiệu Viettel.
        header_fill = PatternFill("solid", fgColor="E6002D")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        thin_gray = Side(style="thin", color="D9DEE7")
        dotted_gray = Side(style="dotted", color="BFC5CE")
        header_border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=thin_gray,
        )
        body_border = Border(
            left=dotted_gray,
            right=dotted_gray,
            top=dotted_gray,
            bottom=dotted_gray,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        worksheet.row_dimensions[1].height = 34

        # Tự tính độ rộng theo tiêu đề và tối đa 500 dòng đầu để xuất nhanh
        # ngay cả khi kết quả có hàng trăm nghìn dòng.
        width_sample_rows = min(max_row, 501)
        date_columns = {
            COL_CREATED,
            COL_DEADLINE,
            COL_FT_FINISHED,
            COL_CD_CLOSED,
        }
        centered_columns = {
            COL_STATUS,
            COL_EMPLOYEE,
            COL_PRIORITY,
            COL_REMAINING,
            COL_STATION,
            COL_SUBSCRIBER,
            COL_SYSTEM,
        }

        column_names = {
            column_index: str(worksheet.cell(1, column_index).value or "")
            for column_index in range(1, max_column + 1)
        }

        for column_index in range(1, max_column + 1):
            column_name = column_names[column_index]
            max_length = len(column_name)

            for row_index in range(2, width_sample_rows + 1):
                value = worksheet.cell(row_index, column_index).value
                if value is None:
                    continue

                if column_name in date_columns:
                    display_length = 19
                else:
                    display_length = len(str(value))

                max_length = max(max_length, display_length)

            # Các cột nội dung dài được giới hạn để bảng không trải quá rộng.
            if column_name in {COL_CONTENT, COL_DESCRIPTION, COL_COORD_GROUP}:
                width = min(max(max_length + 2, 22), 48)
            elif column_name in date_columns:
                width = 21
            elif column_name == COL_WORK_CODE:
                width = min(max(max_length + 2, 22), 35)
            else:
                width = min(max(max_length + 2, 12), 28)

            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        # Định dạng ngày giờ, số và căn lề.
        for column_index, column_name in column_names.items():
            if column_name in date_columns:
                for row_index in range(2, max_row + 1):
                    cell = worksheet.cell(row_index, column_index)
                    if cell.value is not None:
                        cell.number_format = "dd/mm/yyyy hh:mm:ss"
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=False,
                        )
            elif column_name == COL_REMAINING:
                for row_index in range(2, max_row + 1):
                    cell = worksheet.cell(row_index, column_index)
                    if cell.value is not None:
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=False,
                        )
            elif column_name in centered_columns:
                for row_index in range(2, max_row + 1):
                    worksheet.cell(row_index, column_index).alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=False,
                    )
            else:
                for row_index in range(2, max_row + 1):
                    worksheet.cell(row_index, column_index).alignment = Alignment(
                        horizontal="left",
                        vertical=(
                            "top"
                            if column_name in {
                                COL_CONTENT,
                                COL_DESCRIPTION,
                                COL_COORD_GROUP,
                            }
                            else "center"
                        ),
                        wrap_text=False,
                    )

        # Kẻ nét chấm quanh toàn bộ ô dữ liệu, không tự ngắt dòng
        # và cố định chiều cao để bảng gọn hơn.
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            ),
            start=2,
        ):
            worksheet.row_dimensions[row_index].height = 22
            for cell in row:
                cell.border = body_border
                cell.alignment = cell.alignment.copy(
                    wrap_text=False,
                    vertical="center",
                )

        # Tô màu mức độ ưu tiên để dễ nhận biết.
        priority_column_index = next(
            (
                index
                for index, name in column_names.items()
                if name == COL_PRIORITY
            ),
            None,
        )
        if priority_column_index is not None:
            priority_fills = {
                "binh thuong": PatternFill("solid", fgColor="E2F0D9"),
                "nghiem trong": PatternFill("solid", fgColor="FFF2CC"),
                "rat nghiem trong": PatternFill("solid", fgColor="F4CCCC"),
            }
            for row_index in range(2, max_row + 1):
                cell = worksheet.cell(row_index, priority_column_index)
                fill = priority_fills.get(normalize_text(cell.value))
                if fill is not None:
                    cell.fill = fill

        # Thiết lập trang in cơ bản.
        worksheet.print_title_rows = "1:1"
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = "landscape"

    return output.getvalue()


def display_result_table(result: pd.DataFrame):
    """Hiển thị tối đa 20 dòng; kết quả lớn hơn chỉ cho xuất Excel."""
    display_limit = 20
    total_rows = len(result)

    available_columns = [
        column
        for column in DISPLAY_COLUMNS
        if column in result.columns
    ]

    if not available_columns:
        available_columns = list(result.columns)

    if total_rows <= display_limit:
        st.dataframe(
            result[available_columns],
            use_container_width=True,
            hide_index=True,
            height=min(440, 38 + max(total_rows, 1) * 35),
        )
    else:
        st.info(
            f"Kết quả có **{total_rows:,} dòng**, vượt quá giới hạn "
            f"hiển thị **{display_limit} dòng**. "
            "Vui lòng xuất file Excel để xem chi tiết."
        )

    st.download_button(
        "Xuất kết quả ra Excel",
        data=dataframe_to_excel(result),
        file_name=(
            f"KetQua_WO_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key=unique_widget_key("download_result"),
    )


def render_horizontal_bar_chart(
    data: pd.DataFrame,
    category_column: str,
    value_column: str,
    title: str,
    height_per_row: int = 38,
    fixed_height: int | None = None,
):
    """Vẽ biểu đồ thanh ngang màu đỏ, có số liệu ở cuối thanh."""
    if data.empty:
        st.info("Không có dữ liệu để vẽ biểu đồ.")
        return

    chart_data = data.copy()
    chart_data[category_column] = chart_data[category_column].astype(str)
    chart_data[value_column] = pd.to_numeric(
        chart_data[value_column], errors="coerce"
    ).fillna(0)

    order = chart_data[category_column].tolist()
    # Cho phép cố định chiều cao để các biểu đồ đặt cạnh nhau luôn bằng nhau.
    chart_height = (
        fixed_height
        if fixed_height is not None
        else max(180, min(720, len(chart_data) * height_per_row + 55))
    )

    bars = (
        alt.Chart(chart_data)
        .mark_bar(cornerRadiusEnd=7, height=22, color="#e31937")
        .encode(
            y=alt.Y(
                f"{category_column}:N",
                sort=order,
                title=None,
                axis=alt.Axis(labelFontSize=14, labelLimit=260, ticks=False),
            ),
            x=alt.X(
                f"{value_column}:Q",
                title=None,
                axis=None,
                scale=alt.Scale(zero=True, nice=True),
            ),
            tooltip=[
                alt.Tooltip(f"{category_column}:N", title=category_column),
                alt.Tooltip(f"{value_column}:Q", title=value_column, format=",.0f"),
            ],
        )
    )

    labels = (
        alt.Chart(chart_data)
        .mark_text(align="left", baseline="middle", dx=8, fontSize=14, color="#202124")
        .encode(
            y=alt.Y(f"{category_column}:N", sort=order),
            x=alt.X(f"{value_column}:Q"),
            text=alt.Text(f"{value_column}:Q", format=",.0f"),
        )
    )

    chart = (
        (bars + labels)
        .properties(title=title, height=chart_height)
        .configure_title(anchor="start", fontSize=18, fontWeight="bold", offset=16)
        .configure_view(stroke="#e5e7eb", cornerRadius=12)
        .configure(background="#ffffff")
    )

    st.altair_chart(chart, use_container_width=True)


def extract_dashboard_days(question: str, default: int = 7) -> int:
    q = normalize_text(question)
    match = re.search(r"(?:trong|gan nhat)?\s*(\d+)\s*ngay", q)
    if match:
        return max(1, min(int(match.group(1)), 60))
    return default


def answer_dashboard(df: pd.DataFrame, question: str):
    """Dashboard WO đóng; ngày đóng được tính theo Thời điểm FT hoàn thành."""
    required = [COL_STATUS, COL_FT_FINISHED]
    missing = [column for column in required if column not in df.columns]
    if missing:
        st.error("Không tìm thấy cột: " + ", ".join(missing))
        return

    days = extract_dashboard_days(question, default=7)
    today = date.today()
    start_date = today - timedelta(days=days - 1)

    status_series = df[COL_STATUS].fillna("").astype(str).str.strip()
    closed_mask = status_series.str.casefold() == CLOSED_STATUS.casefold()
    finished_dates = df[COL_FT_FINISHED]

    recent_closed = df[
        closed_mask
        & finished_dates.notna()
        & (finished_dates.dt.date >= start_date)
        & (finished_dates.dt.date <= today)
    ].copy()

    total_wo = len(df)
    total_closed = int(closed_mask.sum())
    total_open = total_wo - total_closed
    latest_finished = finished_dates.max()

    st.markdown(
        f"### 📊 Toàn trung tâm — {days} ngày gần nhất: "
        f"đóng <span style='color:#d90429'>{len(recent_closed):,} WO</span>.",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"**Tổng số WO trong file:** <span style='color:#d90429'>{total_wo:,}</span> "
        f"(đã đóng {total_closed:,}, chưa đóng {total_open:,}).",
        unsafe_allow_html=True,
    )
    if pd.notna(latest_finished):
        st.caption(
            "Lần FT hoàn thành gần nhất trong dữ liệu: "
            + latest_finished.strftime("%d/%m/%Y %H:%M")
        )

    status_statistics = (
        status_series.replace("", "Không xác định")
        .value_counts()
        .rename_axis(COL_STATUS)
        .reset_index(name="Số lượng WO")
    )
    render_horizontal_bar_chart(
        status_statistics,
        COL_STATUS,
        "Số lượng WO",
        "Toàn bộ WO theo trạng thái",
    )

    date_index = pd.date_range(start=start_date, end=today, freq="D")
    daily_counts = (
        recent_closed.assign(_ngay=recent_closed[COL_FT_FINISHED].dt.normalize())
        .groupby("_ngay")
        .size()
        .reindex(date_index, fill_value=0)
        .rename("Số WO đóng")
        .reset_index()
        .rename(columns={"index": "Ngày"})
    )
    daily_counts["Ngày"] = daily_counts["Ngày"].dt.strftime("%d/%m")

    render_horizontal_bar_chart(
        daily_counts,
        "Ngày",
        "Số WO đóng",
        f"WO đã đóng theo ngày ({days} ngày gần nhất) — toàn trung tâm",
    )

    st.download_button(
        "Xuất WO đóng trong kỳ ra Excel",
        data=dataframe_to_excel(recent_closed),
        file_name=f"WO_dong_{days}_ngay.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=unique_widget_key("download_dashboard"),
    )


def answer_top(df: pd.DataFrame, question: str):
    if COL_EMPLOYEE not in df.columns:
        st.error("File không có cột Nhân viên thực hiện.")
        return

    # Top mặc định là WO tồn, trừ khi câu hỏi yêu cầu trạng thái cụ thể.
    filtered = df
    requested_status = extract_status(question)

    if requested_status is None and COL_STATUS in filtered.columns:
        status_series = (
            filtered[COL_STATUS]
            .fillna("")
            .astype(str)
            .str.strip()
        )
        filtered = filtered[
            status_series.str.casefold() != CLOSED_STATUS.casefold()
        ]

    # Vẫn áp dụng trạng thái, ngày, quá hạn, trạm... trong câu hỏi.
    filtered, filters = apply_filters(filtered, question)

    top_number = extract_top_number(question)

    statistics = (
        filtered[
            filtered[COL_EMPLOYEE].notna()
            & filtered[COL_EMPLOYEE].astype(str).str.strip().ne("")
        ]
        .groupby(COL_EMPLOYEE)
        .size()
        .reset_index(name="Số lượng WO")
        .sort_values("Số lượng WO", ascending=False)
        .head(top_number)
    )

    st.success(
        f"Top {len(statistics)} nhân viên theo số lượng WO."
    )

    if filters:
        st.caption("Điều kiện: " + " • ".join(filters))

    st.dataframe(
        statistics,
        use_container_width=True,
        hide_index=True,
    )

    if not statistics.empty:
        render_horizontal_bar_chart(
            statistics,
            COL_EMPLOYEE,
            "Số lượng WO",
            "Xếp hạng nhân viên theo số lượng WO",
        )

    st.download_button(
        "Xuất thống kê ra Excel",
        data=dataframe_to_excel(statistics),
        file_name="Top_nhan_vien_WO.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key=unique_widget_key("download_top"),
    )


def answer_statistics(df: pd.DataFrame, question: str):
    result, filters = apply_filters(df, question)
    q = normalize_text(question)

    asks_system = any(phrase in q for phrase in [
        "he thong",
        "ma he thong",
        "theo he thong",
        "theo ma he thong",
    ])
    asks_employee = any(phrase in q for phrase in [
        "nhan vien",
        "tung nhan vien",
        "moi nhan vien",
        "theo ft",
        "theo nhan vien",
    ])
    asks_coord_group = any(phrase in q for phrase in [
        "nhom dieu phoi",
        "theo nhom dieu phoi",
        "tung nhom dieu phoi",
        "moi nhom dieu phoi",
        "nhom dp",
    ])

    # Hỗ trợ tổng hợp theo một hoặc nhiều chiều.
    if asks_coord_group and asks_system and asks_employee:
        group_columns = [COL_COORD_GROUP, COL_SYSTEM, COL_EMPLOYEE]
    elif asks_coord_group and asks_employee:
        group_columns = [COL_COORD_GROUP, COL_EMPLOYEE]
    elif asks_coord_group and asks_system:
        group_columns = [COL_COORD_GROUP, COL_SYSTEM]
    elif asks_system and asks_employee:
        group_columns = [COL_SYSTEM, COL_EMPLOYEE]
    elif asks_coord_group:
        group_columns = [COL_COORD_GROUP]
    elif asks_system:
        group_columns = [COL_SYSTEM]
    elif asks_employee:
        group_columns = [COL_EMPLOYEE]
    elif "tram" in q:
        group_columns = [COL_STATION]
    elif "loai cong viec" in q:
        group_columns = [COL_WORK_TYPE]
    else:
        group_columns = [COL_STATUS]

    missing_columns = [column for column in group_columns if column not in result.columns]
    if missing_columns:
        st.error("Không tìm thấy cột: " + ", ".join(missing_columns))
        return

    valid_mask = pd.Series(True, index=result.index)
    for column in group_columns:
        valid_mask &= (
            result[column].notna()
            & result[column].astype(str).str.strip().ne("")
        )

    statistics = (
        result[valid_mask]
        .groupby(group_columns, dropna=False)
        .size()
        .reset_index(name="Số lượng WO")
        .sort_values("Số lượng WO", ascending=False)
    )

    group_label = " + ".join(group_columns)
    st.success(f"Tổng hợp {len(result):,} WO theo {group_label}.")

    if filters:
        st.caption("Điều kiện: " + " • ".join(filters))

    st.dataframe(
        statistics,
        use_container_width=True,
        hide_index=True,
    )

    # Biểu đồ thanh ngang màu đỏ cho thống kê một chiều.
    if not statistics.empty and len(group_columns) == 1:
        render_horizontal_bar_chart(
            statistics.head(30),
            group_columns[0],
            "Số lượng WO",
            f"Tổng hợp WO theo {group_columns[0]}",
        )

    asks_detail_list = any(phrase in q for phrase in [
        "danh sach",
        "chi tiet",
        "liet ke",
    ])
    if asks_detail_list and not result.empty:
        st.markdown("#### Danh sách chi tiết WO")
        display_result_table(result)

    st.download_button(
        "Xuất thống kê ra Excel",
        data=dataframe_to_excel(statistics),
        file_name="Thong_ke_WO.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key=unique_widget_key("download_statistics"),
    )


def render_help_message() -> None:
    """Hiển thị hướng dẫn và cho phép bấm câu hỏi để gửi trực tiếp cho BOT."""
    st.markdown(
        """
# 🤖 HƯỚNG DẪN SỬ DỤNG TRỢ LÝ TRA CỨU WO

Bạn có thể **bấm trực tiếp vào một câu hỏi** bên dưới để BOT tra cứu ngay.
        """
    )

    help_sections = [
        ("📊 Hỏi tổng hợp WO", [
            "Tổng WO",
            "Tổng WO tồn",
            "Tổng WO của nhóm điều phối Ia Grai",
            "Tổng WO tồn của nhóm điều phối Ia Grai",
            "Chi tiết WO của nhóm điều phối Ia Grai",
            "Chi tiết WO tồn của nhóm điều phối Ia Grai",
            "WO tồn theo trạng thái",
            "WO tồn theo mã hệ thống",
            "Chi tiết WO tồn theo mã hệ thống",
            "Có bao nhiêu WO đang thực hiện?",
            "WO đóng hôm nay là bao nhiêu?",
            "Có bao nhiêu WO hệ thống CC_SCVT?",
            "Có bao nhiêu WO tồn mã hệ thống CC_SCVT?",
            "Có bao nhiêu WO quá hạn hệ thống CC_SCVT?",
            "Có bao nhiêu WO còn dưới 4 giờ?",
        ]),
        ("🔢 Tra cứu số lượng", [
            "Có bao nhiêu WO tồn?",
            "Có bao nhiêu WO quá hạn?",
            "Có bao nhiêu WO đang thực hiện?",
            "WO đóng hôm nay là bao nhiêu?",
            "Có bao nhiêu WO hệ thống CC_SCVT?",
            "Có bao nhiêu WO tồn mã hệ thống CC_SCVT?",
            "Có bao nhiêu WO quá hạn hệ thống CC_SCVT?",
            "Có bao nhiêu WO còn dưới 4 giờ?",
        ]),
        ("👨‍🔧 Tra cứu nhân viên", [
            "WO của binhnt59",
            "WO tồn của binhnt59",
            "WO quá hạn của binhnt59",
            "WO đang thực hiện của binhnt59",
            "WO hệ thống CC_SCVT của binhnt59",
            "WO đóng của binhnt59 trong 5 ngày gần nhất",
            "Top 10 FT tồn nhiều nhất",
            "Top 10 FT quá hạn nhiều nhất",
            "Thống kê WO theo nhân viên",
        ]),
        ("📌 Tra cứu trạng thái", [
            "Danh sách WO FT đang thực hiện",
            "Danh sách WO FT hoàn thành",
            "Danh sách WO chờ CD tiếp nhận",
            "Danh sách WO CD đã tiếp nhận",
            "Danh sách WO FT từ chối",
            "Danh sách WO CD từ chối",
            "Danh sách WO đã giao FT",
            "Thống kê theo trạng thái",
            "WO FT hoàn thành hôm nay",
        ]),
        ("💻 Tra cứu hệ thống", [
            "WO hệ thống CC_SCVT",
            "WO tồn mã hệ thống CC_SCVT",
            "WO tồn hệ thống ICMS",
            "WO quá hạn hệ thống CC_SCVT",
            "WO tồn mã hệ thống CC_SCVT quá hạn",
            "WO tồn mã hệ thống CC_SCVT quá hạn trong ngày mai",
            "WO tồn mã hệ thống CC_SCVT quá hạn trong 2 ngày tiếp theo",
            "Thống kê WO theo hệ thống",
            "Tổng hợp WO theo hệ thống",
            "Tổng hợp WO theo hệ thống và nhân viên",
        ]),
        ("📍 Tra cứu nhóm điều phối", [
            "WO của nhóm điều phối Ia Grai",
            "WO tồn của nhóm điều phối Ia Grai",
            "WO quá hạn của nhóm điều phối Ia Grai",
            "Chi tiết WO của nhóm điều phối Ia Grai",
            "Chi tiết WO tồn của nhóm điều phối Ia Grai",
            "Chi tiết WO tồn quá hạn của nhóm điều phối Ia Grai",
            "Chi tiết WO quá hạn của nhóm điều phối Ia Grai",
            "Tổng hợp WO theo nhóm điều phối",
            "WO tồn của từng nhân viên theo nhóm điều phối",
            "Tổng hợp tồn WO theo nhóm điều phối và nhân viên",
            "Tổng hợp tồn WO theo nhóm điều phối, hệ thống và nhân viên",
        ]),
        ("🔎 Tra cứu mã WO, trạm và thuê bao", [
            "Tra cứu WO WO_123456",
            "Thông tin WO WO_123456",
            "WO trạm GLI001",
            "WO tồn trạm GLI001",
            "WO quá hạn trạm GLI001",
            "WO thuê bao 0961234567",
        ]),
        ("📅 Tra cứu theo thời gian", [
            "WO tạo hôm nay",
            "WO đóng hôm nay",
            "WO đóng hôm qua",
            "WO trong 7 ngày gần nhất",
            "WO trong 30 ngày gần nhất",
            "WO tháng này",
            "WO tạo ngày 19/07/2026",
            "WO đóng ngày 19/07/2026",
            "WO quá hạn ngày 19/07/2026",
            "WO tồn mã hệ thống ICMS quá hạn trong tháng này",
            "WO hệ thống CC_SCVT quá hạn tháng hiện tại",
            "WO quá hạn trong ngày mai",
            "WO quá hạn trong 2 ngày tiếp theo",
            "WO còn dưới 4 giờ",
            "WO còn dưới 8 giờ",
        ]),
        ("📝 Tra cứu theo nội dung công việc", [
            "WO có nội dung siet_kc",
            "WO tồn có nội dung siet_kc",
            "WO quá hạn có nội dung siet_kc",
            "WO tồn có nội dung siet_kc quá hạn trong tháng này",
            "WO tồn có nội dung siet_kc sẽ quá hạn trong tháng này",
        ]),
        ("📂 Tra cứu theo loại công việc", [
            "WO loại công việc ICMS_Kiểm tra",
            "WO tồn loại công việc ICMS_Kiểm tra",
            "WO loại công việc ICMS_Kiểm tra được tạo trong tháng 7",
            "WO loại công việc ICMS_Kiểm tra tháng này",
            "WO loại công việc ICMS_Kiểm tra quá hạn",
            "WO loại công việc ICMS_Kiểm tra của binhnt59",
            "WO loại công việc ICMS_Kiểm tra hệ thống CC_SCVT",
        ]),
        ("📈 Thống kê", [
            "Thống kê WO theo trạng thái",
            "Thống kê WO theo hệ thống",
            "Thống kê WO theo nhân viên",
            "Thống kê WO theo nhóm điều phối",
            "Thống kê WO theo trạm",
            "Thống kê WO theo loại công việc",
            "Tổng hợp WO theo hệ thống và nhân viên",
            "Tổng hợp WO theo nhóm điều phối và nhân viên",
            "Tổng hợp WO theo nhóm điều phối, hệ thống và nhân viên",
        ]),
        ("📊 Biểu đồ", [
            "Vẽ biểu đồ đóng WO 7 ngày gần nhất",
            "Vẽ biểu đồ đóng WO 30 ngày",
            "Dashboard",
        ]),
        ("🔀 Có thể kết hợp nhiều điều kiện", [
            "WO tồn hệ thống CC_SCVT của binhnt59",
            "WO quá hạn của binhnt59 trạm GLI001",
            "WO tồn hệ thống CC_SCVT còn dưới 4 giờ",
            "WO hệ thống CC_SCVT trong 7 ngày gần nhất",
            "WO nhóm điều phối VCC_GLI_005_IGI hệ thống CC_SCVT",
            "WO nội dung mất điện của binhnt59",
            "Top 10 FT quá hạn hệ thống CC_SCVT",
            "Thống kê WO tồn hệ thống CC_SCVT theo nhân viên",
            "WO tồn có nội dung siet_kc sẽ quá hạn trong tháng này",
            "WO loại công việc ICMS_Kiểm tra được tạo trong tháng 7",
        ]),
    ]

    st.markdown(
        """
        <style>
        /* Danh sách câu hỏi Help: gọn, thẳng hàng và căn hoàn toàn bên trái. */
        [class*="st-key-help_question_"] {
            margin:0 !important;
            padding:0 !important;
        }

        [class*="st-key-help_question_"] [data-testid="stButton"] {
            margin:0 !important;
            padding:0 !important;
        }

        /* Xóa khoảng cách mặc định giữa các nút trong từng nhóm Help. */
        [class*="st-key-help_question_group_"] [data-testid="stVerticalBlock"] {
            gap:0 !important;
            row-gap:0 !important;
        }

        [class*="st-key-help_question_group_"] [data-testid="stElementContainer"] {
            margin:0 !important;
            padding:0 !important;
        }

        [class*="st-key-help_question_"] button {
            width:100% !important;
            min-height:24px !important;
            height:24px !important;
            margin:0 !important;
            padding:1px 8px !important;
            display:flex !important;
            align-items:center !important;
            justify-content:flex-start !important;
            text-align:left !important;
            border:0 !important;
            border-radius:5px !important;
            background:transparent !important;
            box-shadow:none !important;
            color:#303643 !important;
            font-size:.91rem !important;
            font-weight:400 !important;
            line-height:1 !important;
        }

        [class*="st-key-help_question_"] button > div,
        [class*="st-key-help_question_"] button p,
        [class*="st-key-help_question_"] button span {
            width:100% !important;
            margin:0 !important;
            padding:0 !important;
            text-align:left !important;
            justify-content:flex-start !important;
            color:#303643 !important;
            line-height:1.25 !important;
        }

        [class*="st-key-help_question_"] button:hover {
            background:#fff1f4 !important;
            color:#e6002d !important;
            transform:none !important;
        }

        [class*="st-key-help_question_"] button:hover p,
        [class*="st-key-help_question_"] button:hover span {
            color:#e6002d !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    question_index = 0
    for section_index, (section_title, questions) in enumerate(help_sections):
        st.markdown(f"---\n\n## {section_title}")
        with st.container(key=f"help_question_group_{section_index}"):
            for question in questions:
                if st.button(
                    f"•  {question}",
                    key=f"help_question_{question_index}",
                    use_container_width=True,
                ):
                    st.session_state["pending_question"] = question
                    st.session_state["show_help_panel"] = False
                    st.rerun()
                question_index += 1

    st.info("Nhập `help` bất cứ lúc nào để mở lại phần hướng dẫn này.")


def extract_future_overdue_days(question: str):
    """Trả về khoảng ngày sắp hết hạn, tính từ ngày mai."""
    q = normalize_text(question)

    if "ngay mai" in q:
        return 1, 1

    match = re.search(r"(\d+)\s*ngay\s*(?:tiep theo|toi|sap toi)", q)
    if match:
        days = max(1, int(match.group(1)))
        return 1, days

    return None


def extract_system(question: str, df: pd.DataFrame):
    """
    Nhận diện hệ thống theo giá trị thật trong cột Hệ thống.

    Chỉ nhận diện khi câu hỏi nêu rõ điều kiện hệ thống, tránh hiểu nhầm
    các từ như FT trong trạng thái "FT từ chối" thành hệ thống WFM-FT.
    """
    if COL_SYSTEM not in df.columns:
        return None

    q = normalize_text(question)

    system_markers = [
        "he thong",
        "ma he thong",
        "thuoc he thong",
        "system",
    ]
    if not any(marker in q for marker in system_markers):
        return None

    values = df[COL_SYSTEM].dropna().astype(str).unique().tolist()

    matches = []
    for value in values:
        normalized_value = normalize_text(value)
        tokens = [
            token for token in re.split(r"[^a-z0-9]+", normalized_value)
            if len(token) >= 2 and token not in {"wo", "he", "thong"}
        ]

        if normalized_value and normalized_value in q:
            matches.append((len(normalized_value), value))
            continue

        if any(re.search(rf"\b{re.escape(token)}\b", q) for token in tokens):
            matches.append((max(map(len, tokens), default=0), value))

    if not matches:
        return None

    matches.sort(key=lambda item: item[0], reverse=True)
    return matches[0][1]


def is_future_overdue_query(question: str) -> bool:
    q = normalize_text(question)
    return (
        extract_future_overdue_days(question) is not None
        and any(phrase in q for phrase in [
            "qua han",
            "sap qua han",
            "gan qua han",
            "het han",
        ])
    )


def filter_future_overdue(df: pd.DataFrame, question: str):
    """
    Lọc WO tồn sẽ hết hạn theo cột Thời điểm yêu cầu kết thúc.

    '2 ngày tiếp theo' = từ ngày mai đến hết ngày kia.
    """
    date_range = extract_future_overdue_days(question)

    if date_range is None or COL_DEADLINE not in df.columns:
        return df.iloc[0:0].copy(), None

    start_offset, end_offset = date_range
    today = pd.Timestamp.now().normalize()
    start_date = today + pd.Timedelta(days=start_offset)
    end_exclusive = today + pd.Timedelta(days=end_offset + 1)

    deadline = pd.to_datetime(df[COL_DEADLINE], errors="coerce")
    status = df[COL_STATUS].fillna("").astype(str).str.strip()

    mask = (
        status.ne(CLOSED_STATUS)
        & deadline.notna()
        & deadline.ge(start_date)
        & deadline.lt(end_exclusive)
    )

    result = df.loc[mask].copy()

    system_name = extract_system(question, df)
    if system_name and COL_SYSTEM in result.columns:
        result = result[
            result[COL_SYSTEM].fillna("").astype(str).eq(str(system_name))
        ]

    return result, (
        start_date.date(),
        (end_exclusive - pd.Timedelta(days=1)).date(),
    )


def filter_current_overdue(df: pd.DataFrame, question: str):
    """Lọc WO tồn đang quá hạn tại thời điểm hiện tại."""
    if COL_DEADLINE not in df.columns:
        return df.iloc[0:0].copy()

    deadline = pd.to_datetime(df[COL_DEADLINE], errors="coerce")
    status = df[COL_STATUS].fillna("").astype(str).str.strip()

    result = df.loc[
        status.ne(CLOSED_STATUS)
        & deadline.notna()
        & deadline.lt(pd.Timestamp.now())
    ].copy()

    system_name = extract_system(question, df)
    if system_name and COL_SYSTEM in result.columns:
        result = result[
            result[COL_SYSTEM].fillna("").astype(str).eq(str(system_name))
        ]

    return result



def filter_overdue_current_month(df: pd.DataFrame, question: str):
    """
    Lọc đúng 3 điều kiện cho câu hỏi dạng:
    "WO tồn có nội dung siet_kc quá hạn trong tháng này".

    Điều kiện:
    1. WO tồn: Trạng thái khác Đóng.
    2. Từ khóa chỉ được tìm trong cột Nội dung công việc.
    3. Sẽ quá hạn trong tháng hiện tại: thời hạn từ hiện tại đến trước
       ngày đầu tiên của tháng kế tiếp.
    """
    required_columns = {COL_STATUS, COL_CONTENT, COL_DEADLINE}
    if not required_columns.issubset(df.columns):
        return df.iloc[0:0].copy(), None, None, None

    now = pd.Timestamp.now()
    month_start = now.normalize().replace(day=1)
    next_month_start = month_start + pd.offsets.MonthBegin(1)
    month_end = next_month_start - pd.Timedelta(seconds=1)

    deadline = pd.to_datetime(df[COL_DEADLINE], errors="coerce")
    status = df[COL_STATUS].fillna("").astype(str).str.strip()

    # Chỉ lấy từ thời điểm hiện tại đến hết tháng này.
    mask = (
        status.ne(CLOSED_STATUS)
        & deadline.notna()
        & deadline.ge(now)
        & deadline.lt(next_month_start)
    )

    content_keyword = extract_content_keyword(question)
    if content_keyword:
        mask &= safe_contains(df[COL_CONTENT], content_keyword)

    result = df.loc[mask].copy()
    return result, now, month_end, content_keyword

def render_special_overdue_query(df: pd.DataFrame, question: str) -> bool:
    """
    Xử lý riêng các câu hỏi quá hạn/sắp quá hạn theo hệ thống.
    Trả về True nếu đã xử lý xong.
    """
    q = normalize_text(question)
    system_name = extract_system(question, df)

    asks_overdue_current_month = (
        "qua han" in q
        and any(phrase in q for phrase in [
            "thang nay",
            "thang hien tai",
            "trong thang hien tai",
        ])
    )

    if asks_overdue_current_month:
        result, start_time, end_time, content_keyword = filter_overdue_current_month(
            df, question
        )

        st.success(
            f"Tìm thấy **{len(result):,} WO tồn sẽ quá hạn trong tháng này**."
        )

        filters = [
            "WO tồn (Trạng thái khác Đóng)",
        ]
        if content_keyword:
            filters.append(
                f"Nội dung công việc chứa: {content_keyword}"
            )
        filters.append(
            f"{COL_DEADLINE} từ {start_time.strftime('%d/%m/%Y %H:%M')} "
            f"đến hết {end_time.strftime('%d/%m/%Y')}"
        )

        st.caption("Điều kiện: " + " • ".join(filters))

        if not result.empty:
            display_result_table(result)
        else:
            st.info("Không có WO phù hợp với đúng 3 điều kiện trên.")

        return True

    if is_future_overdue_query(question):
        result, date_range = filter_future_overdue(df, question)
        start_date, end_date = date_range

        st.success(f"Tìm thấy **{len(result):,} WO**.")

        filters = [
            "WO tồn (Trạng thái khác Đóng)",
            (
                f"Thời điểm yêu cầu kết thúc từ "
                f"{start_date.strftime('%d/%m/%Y')} đến "
                f"{end_date.strftime('%d/%m/%Y')}"
            ),
        ]
        if system_name:
            filters.insert(1, f"Hệ thống: {system_name}")

        st.caption("Điều kiện: " + " • ".join(filters))

        if not result.empty:
            display_result_table(result)
        else:
            st.info("Không có WO phù hợp với điều kiện.")

        return True

    asks_current_overdue = (
        "qua han" in q
        and any(phrase in q for phrase in [
            "dang qua han",
            "wo ton qua han",
            "ton theo ma he thong",
            "theo ma he thong dang qua han",
        ])
    )

    if asks_current_overdue and (
        system_name is not None
        or "theo ma he thong" in q
        or "theo he thong" in q
    ):
        result = filter_current_overdue(df, question)

        if system_name is None and COL_SYSTEM in result.columns:
            summary = (
                result.assign(
                    **{
                        COL_SYSTEM: result[COL_SYSTEM]
                        .fillna("Không xác định")
                        .astype(str)
                    }
                )
                .groupby(COL_SYSTEM, observed=True)
                .size()
                .reset_index(name="Số lượng WO quá hạn")
                .sort_values("Số lượng WO quá hạn", ascending=False)
            )

            st.success(f"Tìm thấy **{len(result):,} WO đang quá hạn**.")
            st.caption(
                "Điều kiện: WO tồn • Thời điểm yêu cầu kết thúc nhỏ hơn hiện tại "
                "• Tổng hợp theo cột Hệ thống"
            )
            st.dataframe(summary, use_container_width=True, hide_index=True)
            return True

        st.success(f"Tìm thấy **{len(result):,} WO đang quá hạn**.")
        caption = (
            "Điều kiện: WO tồn • Đang quá hạn theo cột "
            f"{COL_DEADLINE}"
        )
        if system_name:
            caption += f" • Hệ thống: {system_name}"

        st.caption(caption)

        if not result.empty:
            display_result_table(result)
        else:
            st.info("Không có WO phù hợp với điều kiện.")

        return True

    return False

def answer_question(df: pd.DataFrame, question: str):
    intent = detect_intent(question)

    # Xử lý trước để câu "2 ngày tiếp theo" không bị hiểu nhầm
    # thành lọc theo Thời điểm tạo trong 2 ngày gần đây.
    if render_special_overdue_query(df, question):
        return

    if intent == "help":
        render_help_message()
        return

    if intent == "dashboard":
        answer_dashboard(df, question)
        return

    if intent == "top":
        answer_top(df, question)
        return

    if intent == "statistics":
        answer_statistics(df, question)
        return

    result, filters = apply_filters(df, question)

    if intent == "count":
        st.success(f"Tìm thấy **{len(result):,} WO**.")

        if filters:
            st.caption("Điều kiện: " + " • ".join(filters))

        if COL_STATUS in result.columns and not result.empty:
            status_statistics = (
                result.groupby(COL_STATUS)
                .size()
                .reset_index(name="Số lượng")
                .sort_values("Số lượng", ascending=False)
            )

            st.dataframe(
                status_statistics,
                use_container_width=True,
                hide_index=True,
            )

        return

    if result.empty:
        st.warning(
            "Không tìm thấy WO phù hợp. "
            "Bạn hãy kiểm tra lại tên nhân viên, mã trạm, "
            "mã WO hoặc điều kiện tìm kiếm."
        )

        if filters:
            st.caption(
                "Bot đã nhận diện: " + " • ".join(filters)
            )

        return

    st.success(f"Tìm thấy **{len(result):,} WO**.")

    if filters:
        st.caption("Điều kiện: " + " • ".join(filters))

    display_result_table(result)



def _normalized_status_series(frame: pd.DataFrame) -> pd.Series:
    if COL_STATUS not in frame.columns:
        return pd.Series("", index=frame.index, dtype=str)
    return frame[COL_STATUS].fillna("").astype(str).str.strip().map(normalize_text)


def _pick_dashboard_date_column(frame: pd.DataFrame) -> str | None:
    for column in [COL_CREATED, COL_FT_FINISHED, COL_DEADLINE]:
        if column in frame.columns and frame[column].notna().any():
            return column
    return None


def _dashboard_kpi_card(
    kpi_key: str,
    icon: str,
    color: str,
    label: str,
    value: int,
    note: str,
    filter_query: str = "",
) -> str:
    """Thẻ KPI mang theo bộ lọc trong URL khi mở phần chi tiết."""
    query = f"kpi={kpi_key}"
    if filter_query:
        query = f"{query}&{filter_query}"
    # Mở rộng viewBox quanh tâm (12, 12) để phần nét vẽ chỉ còn 50%.
    # Cách này tác động trực tiếp vào SVG, không phụ thuộc CSS của trình duyệt.
    compact_icon = icon.replace(
        'viewBox="0 0 24 24"',
        'viewBox="-12 -12 48 48"',
        1,
    )
    return (
        f'<a class="kpi-card-link" href="?{escape(query, quote=True)}" target="_self" '
        f'aria-label="Xem chi tiết {label}">'
        f'<div class="kpi-card"><div class="kpi-icon {color}">{compact_icon}</div>'
        f'<div><div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value:,}</div>'
        f'<div class="kpi-note">{note}</div></div></div>'
        f'</a>'
    )



def render_management_dashboard_sql(
    where_sql: str = "",
    params=None,
    filter_query: str = "",
) -> None:
    """Dashboard truy vấn trực tiếp DuckDB, không tải 150.000 dòng vào Pandas."""
    params = params or []
    source = parquet_sql_source()
    now = datetime.now()

    summary_sql = f"""
        SELECT
            COUNT(*) AS total,

            -- Đã đóng: so khớp chính xác trạng thái tiếng Việt.
            SUM(
                CASE
                    WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) = ?
                    THEN 1 ELSE 0
                END
            ) AS closed,

            -- Đang thực hiện: gộp Đã giao FT và FT Đang thực hiện.
            SUM(
                CASE
                    WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) IN (?, ?)
                    THEN 1 ELSE 0
                END
            ) AS doing,

            -- Chờ tiếp nhận: chỉ tính đúng trạng thái Chờ CD tiếp nhận.
            SUM(
                CASE
                    WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) = ?
                    THEN 1 ELSE 0
                END
            ) AS waiting,

            -- Quá hạn SLA chỉ tính WO chưa đóng.
            SUM(
                CASE
                    WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                         AND TRY_CAST("{COL_REMAINING}" AS DOUBLE) < 0
                    THEN 1 ELSE 0
                END
            ) AS overdue,

            -- Năm nhóm cảnh báo theo hệ thống và hạn yêu cầu kết thúc.
            SUM(
                CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = 'ICMS'
                          AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) >= CURRENT_TIMESTAMP
                          AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP + INTERVAL '2 days'
                     THEN 1 ELSE 0 END
            ) AS icms_due_2d,
            SUM(
                CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = 'CC_SCVT'
                          AND CAST(TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) AS DATE) =
                              CAST(CURRENT_DATE + INTERVAL '1 day' AS DATE)
                     THEN 1 ELSE 0 END
            ) AS cc_scvt_due_tomorrow,
            SUM(
                CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = 'ICMS'
                          AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP - INTERVAL '3 days'
                     THEN 1 ELSE 0 END
            ) AS icms_overdue_3d,
            SUM(
                CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = 'CC_SCVT'
                          AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP
                     THEN 1 ELSE 0 END
            ) AS cc_scvt_overdue,
            SUM(
                CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = 'TT'
                          AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP
                     THEN 1 ELSE 0 END
            ) AS tt_overdue,

            -- Năm module nghiệp vụ theo Loại công việc, không tính WO đã đóng.
            SUM(CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?
                     THEN 1 ELSE 0 END) AS maintenance_nt,
            SUM(CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?
                     THEN 1 ELSE 0 END) AS patrol_tc,
            SUM(CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?
                     THEN 1 ELSE 0 END) AS maintenance_thc,
            SUM(CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?
                     THEN 1 ELSE 0 END) AS reinforce_html,
            SUM(CASE WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
                          AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?
                     THEN 1 ELSE 0 END) AS maintenance
        FROM {source}
        {where_sql}
    """
    row = get_duckdb_connection().execute(
        summary_sql,
        [
            CLOSED_STATUS,
            "Đã giao FT",
            "FT Đang thực hiện",
            "Chờ CD tiếp nhận",
            CLOSED_STATUS,
            CLOSED_STATUS,
            CLOSED_STATUS,
            CLOSED_STATUS,
            CLOSED_STATUS,
            CLOSED_STATUS,
            CLOSED_STATUS, "ICMS_Bảo dưỡng nhà trạm",
            CLOSED_STATUS, "ICMS_Tuần tra tuyến cáp",
            CLOSED_STATUS, "ICMS_Bảo dưỡng THC",
            CLOSED_STATUS, "Báo hỏng hạ tầng mạng lưới",
            CLOSED_STATUS, "Bảo dưỡng cứng cơ điện điều hòa, máy phát điện, thông gió lọc bụi ICMS",
        ] + list(params),
    ).fetchone()

    (
        total,
        closed,
        doing,
        waiting,
        overdue,
        icms_due_2d,
        cc_scvt_due_tomorrow,
        icms_overdue_3d,
        cc_scvt_overdue,
        tt_overdue,
        maintenance_nt,
        patrol_tc,
        maintenance_thc,
        reinforce_html,
        maintenance,
    ) = [int(value or 0) for value in row]

    st.markdown(
        f"""
        <div class="dash-topbar">
            <div class="dash-title-wrap">   
                <div>
                    <h2>DASHBOARD QUẢN LÝ WORK ORDER (WO)</h2>
                    <p>Theo dõi · Giám sát · Điều hành · Nâng cao chất lượng dịch vụ</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Giữ nguyên giao diện KPI ban đầu. Mỗi thẻ được bọc bằng liên kết
    # query parameter nên toàn bộ thẻ bấm được mà không làm mất icon/định dạng.
    if "dashboard_selected_kpi" not in st.session_state:
        st.session_state["dashboard_selected_kpi"] = None

    query_kpi = st.query_params.get("kpi")
    if isinstance(query_kpi, list):
        query_kpi = query_kpi[0] if query_kpi else None
    if query_kpi in {
        "total", "closed", "doing", "waiting", "overdue",
        "icms_due_2d", "cc_scvt_due_tomorrow", "icms_overdue_3d",
        "cc_scvt_overdue", "tt_overdue",
        "maintenance_nt", "patrol_tc", "maintenance_thc",
        "reinforce_html", "maintenance",
    }:
        st.session_state["dashboard_selected_kpi"] = query_kpi

    st.markdown(
        """
        <style>
        /* ===== Giữ hàng KPI sát thanh tiêu đề ===== */
        .st-key-dashboard_kpi_clickable {
            /* Thu hẹp khoảng cách với thanh tiêu đề còn khoảng một nửa. */
            margin-top:-10px !important;
            margin-bottom:1px !important;
            padding:0 !important;
        }

        .st-key-dashboard_kpi_clickable > div,
        .st-key-dashboard_kpi_clickable [data-testid="stVerticalBlock"],
        .st-key-dashboard_kpi_clickable [data-testid="stVerticalBlockBorderWrapper"] {
            margin:0 !important;
            padding:0 !important;
            gap:0 !important;
        }

        .st-key-dashboard_kpi_clickable [data-testid="stHorizontalBlock"] {
            gap:10px !important;
            margin:0 !important;
            padding:0 !important;
            align-items:stretch !important;
        }

        .st-key-dashboard_kpi_clickable [data-testid="column"] {
            min-width:0 !important;
            margin:0 !important;
            padding:0 !important;
        }

        .st-key-dashboard_kpi_clickable [data-testid="stMarkdownContainer"],
        .st-key-dashboard_kpi_clickable [data-testid="stMarkdownContainer"] > p {
            margin:0 !important;
            padding:0 !important;
        }

        /* ===== Hàng thẻ cảnh báo thứ hai ===== */
        .st-key-dashboard_alert_kpi,
        .st-key-dashboard_work_type_kpi {
            margin-top:3px !important;
            margin-bottom:2px !important;
            padding:0 !important;
        }

        .st-key-dashboard_alert_kpi > div,
        .st-key-dashboard_work_type_kpi > div,
        .st-key-dashboard_alert_kpi [data-testid="stVerticalBlock"],
        .st-key-dashboard_work_type_kpi [data-testid="stVerticalBlock"],
        .st-key-dashboard_alert_kpi [data-testid="stVerticalBlockBorderWrapper"],
        .st-key-dashboard_work_type_kpi [data-testid="stVerticalBlockBorderWrapper"] {
            margin:0 !important;
            padding:0 !important;
            gap:0 !important;
        }

        .st-key-dashboard_alert_kpi [data-testid="stHorizontalBlock"],
        .st-key-dashboard_work_type_kpi [data-testid="stHorizontalBlock"] {
            gap:10px !important;
            margin:0 !important;
            padding:0 !important;
            align-items:stretch !important;
        }

        .st-key-dashboard_alert_kpi [data-testid="column"],
        .st-key-dashboard_work_type_kpi [data-testid="column"],
        .st-key-dashboard_alert_kpi [data-testid="stMarkdownContainer"],
        .st-key-dashboard_work_type_kpi [data-testid="stMarkdownContainer"],
        .st-key-dashboard_alert_kpi [data-testid="stMarkdownContainer"] > p,
        .st-key-dashboard_work_type_kpi [data-testid="stMarkdownContainer"] > p {
            min-width:0 !important;
            margin:0 !important;
            padding:0 !important;
        }

        .st-key-dashboard_alert_kpi .kpi-label,
        .st-key-dashboard_work_type_kpi .kpi-label {
            white-space:nowrap !important;
            font-size:.85rem !important;
        }

        /* ===== Liên kết bao toàn bộ thẻ KPI ===== */
        .kpi-card-link {
            display:block !important;
            width:100% !important;
            height:100% !important;
            color:inherit !important;
            text-decoration:none !important;
            cursor:pointer !important;
        }

        .kpi-card-link:visited,
        .kpi-card-link:hover,
        .kpi-card-link:active {
            color:inherit !important;
            text-decoration:none !important;
        }

        /* ===== Giữ nguyên bố cục thẻ KPI cũ ===== */
        .kpi-card-link .kpi-card {
            width:100% !important;
            min-height:84px !important;
            box-sizing:border-box !important;

            display:flex !important;
            align-items:center !important;
            justify-content:flex-start !important;
            gap:12px !important;

            margin:0 !important;
            padding:12px 13px !important;

            transition:
                transform .16s ease,
                box-shadow .16s ease,
                border-color .16s ease;
        }

        /* ===== Căn icon đúng giữa ===== */
        .kpi-card-link .kpi-icon {
            width:44px !important;
            height:44px !important;
            min-width:44px !important;
            flex:0 0 44px !important;

            display:grid !important;
            place-items:center !important;

            margin:0 !important;
            padding:0 !important;
            line-height:1 !important;
        }

        .kpi-card-link .kpi-icon svg {
            display:block !important;
            width:24px !important;
            height:24px !important;
            margin:0 !important;
            overflow:visible;
            fill:none;
            stroke:currentColor;
            stroke-width:5.2;
            stroke-linecap:round;
            stroke-linejoin:round;
        }

        /* ===== Căn khối chữ bên phải icon ===== */
        .kpi-card-link .kpi-card > div:last-child {
            min-width:0 !important;
            display:flex !important;
            flex-direction:column !important;
            justify-content:center !important;
            align-items:flex-start !important;
        }

        .kpi-card-link .kpi-label,
        .kpi-card-link .kpi-value,
        .kpi-card-link .kpi-note {
            margin:0 !important;
            padding:0 !important;
            text-align:left !important;
        }

        .kpi-card-link .kpi-value {
            margin-top:4px !important;
        }

        .kpi-card-link .kpi-note {
            margin-top:5px !important;
        }

        /* ===== Hiệu ứng khi rê chuột ===== */
        .kpi-card-link:hover .kpi-card {
            transform:translateY(-2px);
            border-color:rgba(230,0,45,.35) !important;
            box-shadow:0 12px 25px rgba(230,0,45,.14) !important;
        }

        .kpi-card-link:focus-visible .kpi-card {
            outline:3px solid rgba(230,0,45,.18);
            outline-offset:2px;
        }

        /* ===== Khung hiển thị chi tiết KPI ===== */
        .st-key-dashboard_kpi_detail {
            margin:2px 0 14px !important;
            padding:14px !important;
            border:1px solid #e5e9f0 !important;
            border-radius:14px !important;
            background:#fff !important;
            box-shadow:0 7px 18px rgba(20,32,51,.05) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Dùng cùng viewBox và nét vẽ để mọi icon có tâm hình học đồng nhất.
    kpi_icons = {
        "total": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><rect x="5" y="4" width="14" height="16" rx="1.5" fill="none" stroke="white"/><path d="M8 8h8M8 12h8M8 16h8" fill="none" stroke="white"/></svg>',
        "closed": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><path d="M5 12.5l4.2 4.2L19 7" fill="none" stroke="white"/></svg>',
        "doing": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><circle cx="12" cy="12" r="8.5" fill="none" stroke="white"/><path d="M12 6.5V12h5.5" fill="none" stroke="white"/></svg>',
        "waiting": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><circle cx="12" cy="12" r="8.5" fill="none" stroke="white"/><circle cx="12" cy="12" r="4.5" fill="none" stroke="white"/></svg>',
        "overdue": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><path d="M12 5.5v8" fill="none" stroke="white"/><circle cx="12" cy="18" r="1.15" fill="white" stroke="none"/></svg>',
        "calendar": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><rect x="4" y="5.5" width="16" height="14" rx="2" fill="none" stroke="white"/><path d="M8 3.5v4M16 3.5v4M4 9.5h16M8 13h3M8 16h5" fill="none" stroke="white"/></svg>',
        "calendar_clock": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><rect x="3.5" y="5" width="17" height="15" rx="2" fill="none" stroke="white"/><path d="M7.5 3v4M16.5 3v4M3.5 9h17" fill="none" stroke="white"/><circle cx="15.5" cy="15" r="3" fill="none" stroke="white"/><path d="M15.5 13.5V15l1 1" fill="none" stroke="white"/></svg>',
        "hourglass": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><path d="M7 3.5h10M7 20.5h10M8 3.5c0 4 1.5 5.5 4 8.5-2.5 3-4 4.5-4 8.5M16 3.5c0 4-1.5 5.5-4 8.5 2.5 3 4 4.5 4 8.5" fill="none" stroke="white"/></svg>',
        "warning": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><path d="M12 3.5L21 20H3L12 3.5zM12 9v5" fill="none" stroke="white"/><circle cx="12" cy="17" r=".9" fill="white" stroke="none"/></svg>',
        "alarm": '<svg viewBox="0 0 24 24" fill="none" stroke="white" aria-hidden="true"><circle cx="12" cy="13" r="7" fill="none" stroke="white"/><path d="M12 9v4l2.5 1.5M7 3.5L3.5 7M17 3.5L20.5 7M8 20l-1.5 1.5M16 20l1.5 1.5" fill="none" stroke="white"/></svg>',
    }
    kpi_items = [
        ("total", kpi_icons["total"], "red", "Tổng số WO", total, "Toàn bộ dữ liệu đang lọc"),
        ("closed", kpi_icons["closed"], "green", "Đã đóng", closed, f"{(closed/total*100 if total else 0):.1f}% tổng WO"),
        ("doing", kpi_icons["doing"], "orange", "Đã giao/FT tiếp nhận", doing, f"{(doing/total*100 if total else 0):.1f}% tổng WO"),
        ("waiting", kpi_icons["waiting"], "blue", "Chờ CD tiếp nhận", waiting, f"{(waiting/total*100 if total else 0):.1f}% tổng WO"),
        ("overdue", kpi_icons["overdue"], "darkred", "Quá hạn SLA", overdue, f"{(overdue/total*100 if total else 0):.1f}% tổng WO"),
    ]

    with st.container(key="dashboard_kpi_clickable"):
        kpi_columns = st.columns(5, gap="small")
        for column, (kpi_key, icon, color, label, value, note) in zip(kpi_columns, kpi_items):
            with column:
                st.markdown(
                    _dashboard_kpi_card(
                        kpi_key, icon, color, label, value, note, filter_query
                    ),
                    unsafe_allow_html=True,
                )

    alert_kpi_items = [
        (
            "icms_due_2d", kpi_icons["calendar_clock"], "orange",
            "ICMS QH 2 ngày tới", icms_due_2d,
            "Hạn trong 48 giờ tới",
        ),
        (
            "cc_scvt_due_tomorrow", kpi_icons["calendar"], "blue",
            "SCVT QH trong ngày mai", cc_scvt_due_tomorrow,
            "Hạn vào ngày kế tiếp",
        ),
        (
            "icms_overdue_3d", kpi_icons["hourglass"], "darkred",
            "ICMS QH > 3 NGÀY", icms_overdue_3d,
            "Đã trễ trên 72 giờ",
        ),
        (
            "cc_scvt_overdue", kpi_icons["warning"], "red",
            "CC_SCVT QUÁ HẠN", cc_scvt_overdue,
            "Hạn kết thúc đã qua",
        ),
        (
            "tt_overdue", kpi_icons["alarm"], "green",
            "WO TT QH", tt_overdue,
            "Hạn kết thúc đã qua",
        ),
    ]

    with st.container(key="dashboard_alert_kpi"):
        alert_columns = st.columns(5, gap="small")
        for column, (kpi_key, icon, color, label, value, note) in zip(
            alert_columns, alert_kpi_items
        ):
            with column:
                st.markdown(
                    _dashboard_kpi_card(
                        kpi_key, icon, color, label, value, note, filter_query
                    ),
                    unsafe_allow_html=True,
                )


    work_type_kpi_items = [
        (
            "maintenance_nt", kpi_icons["total"], "red",
            "BẢO DƯỠNG NT", maintenance_nt,
            "ICMS_Bảo dưỡng nhà trạm · Chưa đóng",
        ),
        (
            "patrol_tc", kpi_icons["calendar"], "blue",
            "TUẦN TRA TC", patrol_tc,
            "ICMS_Tuần tra tuyến cáp · Chưa đóng",
        ),
        (
            "maintenance_thc", kpi_icons["doing"], "orange",
            "BẢO DƯỠNG THC", maintenance_thc,
            "ICMS_Bảo dưỡng THC · Chưa đóng",
        ),
        (
            "reinforce_html", kpi_icons["warning"], "darkred",
            "CỦNG CỐ HTML", reinforce_html,
            "Báo hỏng hạ tầng mạng lưới · Chưa đóng",
        ),
        (
            "maintenance", kpi_icons["hourglass"], "green",
            "MAINTENANCE", maintenance,
            "Bảo dưỡng cơ điện, điều hòa... · Chưa đóng",
        ),
    ]

    with st.container(key="dashboard_work_type_kpi"):
        work_type_columns = st.columns(5, gap="small")
        for column, (kpi_key, icon, color, label, value, note) in zip(
            work_type_columns, work_type_kpi_items
        ):
            with column:
                st.markdown(
                    _dashboard_kpi_card(
                        kpi_key, icon, color, label, value, note, filter_query
                    ),
                    unsafe_allow_html=True,
                )

    selected_kpi = st.session_state.get("dashboard_selected_kpi")
    if selected_kpi:
        kpi_conditions = {
            "total": ("1=1", [], "TỔNG SỐ WO"),
            "closed": (f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) = ?', [CLOSED_STATUS], "WO ĐÃ ĐÓNG"),
            "doing": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) IN (?, ?)',
                ["Đã giao FT", "FT Đang thực hiện"],
                "WO ĐÃ GIAO/FT TIẾP NHẬN",
            ),
            "waiting": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) = ?',
                ["Chờ CD tiếp nhận"],
                "WO CHỜ CD TIẾP NHẬN",
            ),
            "overdue": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND (TRY_CAST("{COL_REMAINING}" AS DOUBLE) < 0 '
                f'OR TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP)',
                [CLOSED_STATUS],
                "WO QUÁ HẠN SLA",
            ),
            "icms_due_2d": (
                f'UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = ? '
                f'AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) >= CURRENT_TIMESTAMP '
                f'AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < '
                f"CURRENT_TIMESTAMP + INTERVAL '2 days'",
                ["ICMS", CLOSED_STATUS],
                "ICMS SẮP QUÁ HẠN TRONG 2 NGÀY",
            ),
            "cc_scvt_due_tomorrow": (
                f'UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = ? '
                f'AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND CAST(TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) AS DATE) = '
                f"CAST(CURRENT_DATE + INTERVAL '1 day' AS DATE)",
                ["CC_SCVT", CLOSED_STATUS],
                "CC_SCVT SẮP QUÁ HẠN NGÀY MAI",
            ),
            "icms_overdue_3d": (
                f'UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = ? '
                f'AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < '
                f"CURRENT_TIMESTAMP - INTERVAL '3 days'",
                ["ICMS", CLOSED_STATUS],
                "ICMS QUÁ HẠN LỚN HƠN 3 NGÀY",
            ),
            "cc_scvt_overdue": (
                f'UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = ? '
                f'AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP',
                ["CC_SCVT", CLOSED_STATUS],
                "CC_SCVT ĐANG QUÁ HẠN",
            ),
            "tt_overdue": (
                f'UPPER(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR))) = ? '
                f'AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP',
                ["TT", CLOSED_STATUS],
                "TT QUÁ HẠN",
            ),

            "maintenance_nt": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?',
                [CLOSED_STATUS, "ICMS_Bảo dưỡng nhà trạm"],
                "BẢO DƯỠNG NT CHƯA ĐÓNG",
            ),
            "patrol_tc": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?',
                [CLOSED_STATUS, "ICMS_Tuần tra tuyến cáp"],
                "TUẦN TRA TC CHƯA ĐÓNG",
            ),
            "maintenance_thc": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?',
                [CLOSED_STATUS, "ICMS_Bảo dưỡng THC"],
                "BẢO DƯỠNG THC CHƯA ĐÓNG",
            ),
            "reinforce_html": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?',
                [CLOSED_STATUS, "Báo hỏng hạ tầng mạng lưới"],
                "CỦNG CỐ HTML CHƯA ĐÓNG",
            ),
            "maintenance": (
                f'TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ? '
                f'AND TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)) = ?',
                [CLOSED_STATUS, "Bảo dưỡng cứng cơ điện điều hòa, máy phát điện, thông gió lọc bụi ICMS"],
                "MAINTENANCE CHƯA ĐÓNG",
            ),
        }
        condition_sql, condition_params, detail_title = kpi_conditions[selected_kpi]
        selected_columns = [column for column in DISPLAY_COLUMNS if column]
        quoted_columns = ", ".join(f'"{column}"' for column in selected_columns)
        kpi_detail_sql = f"""
            WITH base AS (
                SELECT *
                FROM {source}
                {where_sql}
            )
            SELECT {quoted_columns}
            FROM base
            WHERE {condition_sql}
            ORDER BY TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) ASC NULLS LAST
        """
        kpi_detail = duckdb_df(kpi_detail_sql, list(params) + condition_params)

        with st.container(key="dashboard_kpi_detail"):
            title_col, close_col = st.columns([8, 1])
            with title_col:
                st.markdown(f"### {detail_title} — {len(kpi_detail):,} dòng")
            with close_col:
                if st.button("✕ Đóng", key="close_dashboard_kpi_detail"):
                    st.session_state["dashboard_selected_kpi"] = None
                    if "kpi" in st.query_params:
                        del st.query_params["kpi"]
                    st.rerun()

            if kpi_detail.empty:
                st.info("Không có dữ liệu phù hợp với ô đã chọn.")
            else:
                st.dataframe(
                    kpi_detail.head(MAX_DISPLAY_ROWS),
                    use_container_width=True,
                    hide_index=True,
                    height=360,
                )
                if len(kpi_detail) > MAX_DISPLAY_ROWS:
                    st.caption(
                        f"Đang hiển thị {MAX_DISPLAY_ROWS:,}/{len(kpi_detail):,} dòng. "
                        "File Excel bên dưới chứa đầy đủ dữ liệu."
                    )
                st.download_button(
                    "⇩ Tải chi tiết ra Excel",
                    data=dataframe_to_excel(kpi_detail),
                    file_name=f"{selected_kpi}_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_dashboard_kpi_{selected_kpi}",
                )

    # Biểu đồ xu hướng dùng đúng mốc thời gian cho từng chỉ tiêu:
    # - Tổng WO, Đang thực hiện, Quá hạn: theo Thời điểm tạo.
    # - Đã đóng: theo Thời điểm FT hoàn thành.
    # Không dùng Thời điểm FT hoàn thành làm trục chung vì sẽ loại bỏ
    # hầu hết WO đang thực hiện/chờ tiếp nhận chưa có thời điểm hoàn thành.
    trend_sql = f"""
        WITH base AS (
            SELECT *
            FROM {source}
            {where_sql}
        ),
        created_daily AS (
            SELECT
                CAST("{COL_CREATED}" AS DATE) AS ngay,
                COUNT(*) AS total_wo,
                SUM(
                    CASE
                        WHEN TRIM(CAST("{COL_STATUS}" AS VARCHAR)) = ?
                        THEN 1 ELSE 0
                    END
                ) AS doing_wo
            FROM base
            WHERE "{COL_CREATED}" IS NOT NULL
            GROUP BY 1
        ),
        overdue_daily AS (
            -- WO quá hạn phải được nhóm theo ngày hết hạn SLA,
            -- không nhóm theo ngày tạo. Nếu nhóm theo ngày tạo thì phần lớn
            -- WO tồn cũ sẽ dồn về các ngày trước và đường biểu đồ gần bằng 0.
            SELECT
                CAST("{COL_DEADLINE}" AS DATE) AS ngay,
                COUNT(*) AS overdue_wo
            FROM base
            WHERE "{COL_DEADLINE}" IS NOT NULL
              AND TRIM(CAST("{COL_STATUS}" AS VARCHAR)) <> ?
              AND (
                    TRY_CAST("{COL_REMAINING}" AS DOUBLE) < 0
                    OR TRY_CAST("{COL_DEADLINE}" AS TIMESTAMP) < CURRENT_TIMESTAMP
                  )
            GROUP BY 1
        ),
        closed_daily AS (
            SELECT
                CAST("{COL_FT_FINISHED}" AS DATE) AS ngay,
                COUNT(*) AS closed_wo
            FROM base
            WHERE "{COL_FT_FINISHED}" IS NOT NULL
            GROUP BY 1
        ),
        all_days AS (
            SELECT ngay FROM created_daily
            UNION
            SELECT ngay FROM overdue_daily
            UNION
            SELECT ngay FROM closed_daily
        )
        SELECT
            d.ngay AS "Ngày",
            COALESCE(c.total_wo, 0) AS "Tổng WO",
            COALESCE(c.doing_wo, 0) AS "Đang thực hiện",
            COALESCE(f.closed_wo, 0) AS "Đã đóng",
            COALESCE(o.overdue_wo, 0) AS "Quá hạn"
        FROM all_days d
        LEFT JOIN created_daily c ON c.ngay = d.ngay
        LEFT JOIN overdue_daily o ON o.ngay = d.ngay
        LEFT JOIN closed_daily f ON f.ngay = d.ngay
        ORDER BY d.ngay DESC
        LIMIT 30
    """
    # Trong câu SQL, các dấu ? của base/{where_sql} xuất hiện trước
    # các dấu ? dùng để so sánh trạng thái trong created_daily.
    # Vì vậy phải truyền dashboard_params trước, rồi mới đến trạng thái.
    trend_params = list(params) + [
        "FT Đang thực hiện",
        CLOSED_STATUS,
    ]

    trend = duckdb_df(
        trend_sql,
        trend_params,
    ).sort_values("Ngày")

    # Năng suất FT trong tháng hiện tại, dựa trên cột Thời điểm FT hoàn thành.
    # Trung bình/ngày = tổng WO FT hoàn thành từ đầu tháng đến hôm nay
    # chia cho số ngày lịch đã trôi qua trong tháng.
    employee_productivity_sql = f"""
        WITH base AS (
            SELECT *
            FROM {source}
            {where_sql}
        )
        SELECT
            CAST("{COL_EMPLOYEE}" AS VARCHAR) AS "{COL_EMPLOYEE}",
            SUM(
                CASE
                    WHEN CAST("{COL_FT_FINISHED}" AS DATE) = CURRENT_DATE
                    THEN 1 ELSE 0
                END
            ) AS "WO hôm nay",
            COUNT(*) AS "WO tháng này",
            ROUND(
                COUNT(*) * 1.0 /
                GREATEST(EXTRACT(DAY FROM CURRENT_DATE), 1),
                2
            ) AS "Trung bình/ngày"
        FROM base
        WHERE "{COL_EMPLOYEE}" IS NOT NULL
          AND TRIM(CAST("{COL_EMPLOYEE}" AS VARCHAR)) <> ''
          AND "{COL_FT_FINISHED}" IS NOT NULL
          AND CAST("{COL_FT_FINISHED}" AS DATE)
              BETWEEN DATE_TRUNC('month', CURRENT_DATE) AND CURRENT_DATE
        GROUP BY 1
        ORDER BY "Trung bình/ngày" DESC, "WO hôm nay" DESC
        LIMIT 15
    """
    employee_productivity = duckdb_df(employee_productivity_sql, params)

    # Hai biểu đồ nằm cùng một hàng, chia đều 50% - 50%.
    col_trend, col_productivity = st.columns(2, gap="small")

    with col_trend:
        with st.container(border=True, height=390):
            st.markdown(
                '<div class="chart-card-title">XU HƯỚNG WO THEO NGÀY '
                '(TẠO / HẾT HẠN / FT HOÀN THÀNH)</div>',
                unsafe_allow_html=True,
            )
            if not trend.empty:
                long_daily = trend.melt(
                    "Ngày",
                    var_name="Chỉ tiêu",
                    value_name="Số lượng",
                )
                chart = (
                    alt.Chart(long_daily)
                    .mark_line(point=True, strokeWidth=2)
                    .encode(
                        x=alt.X(
                            "Ngày:T",
                            title=None,
                            axis=alt.Axis(format="%d", labelAngle=0),
                        ),
                        y=alt.Y("Số lượng:Q", title=None),
                        color=alt.Color(
                            "Chỉ tiêu:N",
                            legend=alt.Legend(orient="top", title=None),
                        ),
                        tooltip=["Ngày:T", "Chỉ tiêu:N", "Số lượng:Q"],
                    )
                    .properties(height=330)
                )
                st.altair_chart(chart, use_container_width=True)
            else:
                st.info("Không có dữ liệu theo thời gian.")

    with col_productivity:
        with st.container(border=True, height=390):
            st.markdown(
                '<div class="chart-card-title">'
                'WO TỒN VÀ WO TỒN QUÁ HẠN THEO NHÓM ĐIỀU PHỐI'
                '</div>',
                unsafe_allow_html=True,
            )

            coord_group_aliases = {
                "Trung tâm Chư Sê (VCC_GLI_004_CSE)": "CSE",
                "Trung tâm Phù Cát (VCC_GLI_010_PCT)": "PCT",
                "Trung tâm Quy Nhơn Đông (VCC_GLI_009_QND)": "QND",
                "Trung tâm Chư Prông (VCC_GLI_003_CPG)": "CPG",
                "Trung tâm Tây Sơn (VCC_GLI_011_TSN)": "TSN",
                "Trung tâm Đak Đoa (VCC_GLI_008_DDA)": "DDA",
                "Trung tâm Hội Phú (VCC_GLI_007_HPU)": "HPU",
                "Trung tâm Bồng Sơn (VCC_GLI_012_BSN)": "BSN",
                "Trung tâm Tuy Phước Tây (VCC_GLI_013_TPT)": "TPT",
                "Trung tâm Ayun Pa (VCC_GLI_002_APA)": "APA",
                "Trung tâm An Khê (VCC_GLI_001_AKE)": "AKE",
                "Trung tâm Kbang (VCC_GLI_006_KBG)": "KBG",
                "Trung tâm Ia Grai (VCC_GLI_005_IGI)": "IGI",
            }

            coord_group_names = list(coord_group_aliases.keys())
            coord_placeholders = ",".join(
                ["?"] * len(coord_group_names)
            )

            coord_chart_sql = f"""
                WITH base AS (
                    SELECT *
                    FROM {source}
                    {where_sql}
                )
                SELECT
                    CAST("{COL_COORD_GROUP}" AS VARCHAR)
                        AS "{COL_COORD_GROUP}",

                    SUM(
                        CASE
                            WHEN LOWER(
                                TRIM(CAST("{COL_STATUS}" AS VARCHAR))
                            ) <> LOWER(?)
                             AND (
                                "{COL_DEADLINE}" IS NULL
                                OR CAST("{COL_DEADLINE}" AS TIMESTAMP)
                                   >= CURRENT_TIMESTAMP
                             )
                            THEN 1 ELSE 0
                        END
                    ) AS "WO tồn",

                    SUM(
                        CASE
                            WHEN LOWER(
                                TRIM(CAST("{COL_STATUS}" AS VARCHAR))
                            ) <> LOWER(?)
                             AND CAST("{COL_DEADLINE}" AS TIMESTAMP)
                                 < CURRENT_TIMESTAMP
                            THEN 1 ELSE 0
                        END
                    ) AS "WO tồn quá hạn"

                FROM base
                WHERE CAST("{COL_COORD_GROUP}" AS VARCHAR)
                      IN ({coord_placeholders})
                GROUP BY 1
            """

            coord_chart_df = duckdb_df(
                coord_chart_sql,
                list(params)
                + [CLOSED_STATUS, CLOSED_STATUS]
                + coord_group_names,
            )

            if not coord_chart_df.empty:
                coord_chart_df["Nhóm"] = (
                    coord_chart_df[COL_COORD_GROUP]
                    .map(coord_group_aliases)
                )

                coord_chart_df["Tổng WO tồn"] = (
                    coord_chart_df["WO tồn"]
                    + coord_chart_df["WO tồn quá hạn"]
                )

                coord_chart_df = (
                    coord_chart_df
                    .sort_values(
                        by=[
                            "Tổng WO tồn",
                            "WO tồn quá hạn",
                            "Nhóm",
                        ],
                        ascending=[False, False, True],
                    )
                    .reset_index(drop=True)
                )

                group_order = coord_chart_df["Nhóm"].tolist()

                coord_long = coord_chart_df.melt(
                    id_vars=[
                        "Nhóm",
                        "Tổng WO tồn",
                    ],
                    value_vars=[
                        "WO tồn",
                        "WO tồn quá hạn",
                    ],
                    var_name="Chỉ tiêu",
                    value_name="Số lượng",
                )

                coord_long["Thứ tự"] = (
                    coord_long["Chỉ tiêu"]
                    .map({
                        "WO tồn": 0,
                        "WO tồn quá hạn": 1,
                    })
                    .astype(int)
                )

                base_coord_chart = alt.Chart(
                    coord_long
                ).encode(
                    x=alt.X(
                        "Nhóm:N",
                        sort=group_order,
                        title=None,
                        axis=alt.Axis(
                            labels=False,
                            ticks=False,
                            domain=False,
                            title=None,
                        ),
                    ),
                    order=alt.Order(
                        "Thứ tự:Q",
                        sort="ascending",
                    ),
                )

                stacked_bars = base_coord_chart.mark_bar(
                    cornerRadiusTopLeft=4,
                    cornerRadiusTopRight=4,
                    size=28,
                ).encode(
                    y=alt.Y(
                        "Số lượng:Q",
                        stack="zero",
                        title=None,
                        axis=alt.Axis(
                            format=",d",
                            tickMinStep=1,
                        ),
                    ),
                    color=alt.Color(
                        "Chỉ tiêu:N",
                        scale=alt.Scale(
                            domain=[
                                "WO tồn",
                                "WO tồn quá hạn",
                            ],
                            range=[
                                "#B5B5B5",
                                "#EE0033",
                            ],
                        ),
                        legend=alt.Legend(
                            orient="top",
                            title=None,
                            direction="horizontal",
                        ),
                    ),
                    tooltip=[
                        alt.Tooltip(
                            "Nhóm:N",
                            title="Nhóm điều phối",
                        ),
                        alt.Tooltip(
                            "Chỉ tiêu:N",
                            title="Chỉ tiêu",
                        ),
                        alt.Tooltip(
                            "Số lượng:Q",
                            title="Số WO",
                            format=",d",
                        ),
                        alt.Tooltip(
                            "Tổng WO tồn:Q",
                            title="Tổng WO tồn",
                            format=",d",
                        ),
                    ],
                )

                # Biểu đồ chính.
                coord_chart = stacked_bars.properties(height=200)

                # Các dòng thông tin bên dưới dùng cùng trục X với biểu đồ,
                # nên luôn căn thẳng chính xác với từng cột.
                group_label_chart = (
                    alt.Chart(coord_chart_df)
                    .mark_text(
                        fontSize=10,
                        fontWeight="bold",
                        color="#344054",
                    )
                    .encode(
                        x=alt.X(
                            "Nhóm:N",
                            sort=group_order,
                            axis=None,
                        ),
                        y=alt.value(10),
                        text=alt.Text("Nhóm:N"),
                    )
                    .properties(height=18)
                )

                wo_ton_label_chart = (
                    alt.Chart(coord_chart_df)
                    .mark_text(
                        fontSize=10,
                        fontWeight="bold",
                        color="#2563eb",
                    )
                    .encode(
                        x=alt.X(
                            "Nhóm:N",
                            sort=group_order,
                            axis=None,
                        ),
                        y=alt.value(10),
                        text=alt.Text(
                            "WO tồn:Q",
                            format=",d",
                        ),
                    )
                    .properties(height=18)
                )

                qua_han_label_chart = (
                    alt.Chart(coord_chart_df)
                    .mark_text(
                        fontSize=10,
                        fontWeight="bold",
                        color="#e6002d",
                    )
                    .encode(
                        x=alt.X(
                            "Nhóm:N",
                            sort=group_order,
                            axis=None,
                        ),
                        y=alt.value(10),
                        text=alt.Text(
                            "WO tồn quá hạn:Q",
                            format=",d",
                        ),
                    )
                    .properties(height=18)
                )

                coord_full_chart = alt.vconcat(
                    coord_chart,
                    group_label_chart,
                    wo_ton_label_chart,
                    qua_han_label_chart,
                    spacing=-1,
                ).resolve_scale(
                    x="shared",
                )

                st.altair_chart(
                    coord_full_chart,
                    use_container_width=True,
                )
            else:
                st.info(
                    "Không có dữ liệu của 13 nhóm điều phối "
                    "phù hợp với bộ lọc hiện tại."
                )

    work_type_sql = f"""
        SELECT
            COALESCE(NULLIF(TRIM(CAST("{COL_WORK_TYPE}" AS VARCHAR)), ''), 'Không xác định') AS "{COL_WORK_TYPE}",
            COUNT(*) AS "Số lượng WO"
        FROM {source}
        {where_sql}
        GROUP BY 1
        ORDER BY 2 DESC
        LIMIT 12
    """
    system_sql = f"""
        SELECT
            COALESCE(NULLIF(TRIM(CAST("{COL_SYSTEM}" AS VARCHAR)), ''), 'Không xác định') AS "{COL_SYSTEM}",
            COUNT(*) AS "Số lượng WO"
        FROM {source}
        {where_sql}
        GROUP BY 1
        ORDER BY 2 DESC
        LIMIT 12
    """
    work_type_df = duckdb_df(work_type_sql, params)
    system_df = duckdb_df(system_sql, params)

    col_work, col_system = st.columns(2, gap="small")
    with col_work:
        with st.container(border=True):
            st.markdown('<div class="chart-card-title">WO THEO LOẠI CÔNG VIỆC</div>', unsafe_allow_html=True)
            if not work_type_df.empty:
                render_horizontal_bar_chart(
                    work_type_df,
                    COL_WORK_TYPE,
                    "Số lượng WO",
                    "",
                    fixed_height=511,
                )

    with col_system:
        with st.container(border=True):
            st.markdown('<div class="chart-card-title">WO THEO HỆ THỐNG</div>', unsafe_allow_html=True)
            if not system_df.empty:
                render_horizontal_bar_chart(
                    system_df,
                    COL_SYSTEM,
                    "Số lượng WO",
                    "",
                    fixed_height=511,
                )



# ============================================================
# GIAO DIỆN
# ============================================================

# Logo Viettel Construction được nhúng trực tiếp vào code.
# Không cần đặt thêm file ảnh trong thư mục chạy ứng dụng.
VIETTEL_LOGO_BASE64 = "UklGRi4aAABXRUJQVlA4WAoAAAAQAAAAKwEAbQAAQUxQSKwRAAAR/4KgbduG3niPiHhE0u2vEyJSEEQkREiEHRGJBCF4GwDaTaJt29ZKQiiCWFBG1BHLYBnLjG3UMQp2pAhBarKz/v/v2MnKTuQ4jvPTdUX0fwLga5tHQ/J1qjn4X6e271KgV01MDSNXKBaLy/NpjART88vFYrGQN6eQllksFovFwqwxZTIdYo6L0yL3tz9xHMcZtitGBPr395HjOM548LSE0yax3xk7juOM+38z02Vu4DHF+ZRIP1PwpILKcG1Ewe3ZKaMduBTo3SemyvyQ411NiTWXQe+mMuOJmN7xlJnpEnO8MmWI6WxOiS2P080oS3c4dD1l5kccsTtV8gNOPfVPkun+U+xPl6EkhBDjegH+g3gnU2V+RG7jsrK9t7Ocgv8kv6fM4DCJMF3/y3TKCNN289+NaSZTqVQqmdBjg4aZ8k0a+RNdT6ZSqVRSjwR1w0ymApOmaeioQjP9twXHzpuq8zbrNsVNmmYC1Wi6YSZTwaZp6KhsGEm10Wq32+3W+9/NHMbBWDx4arZ9m7t6YvGq1W63263qmq4EjczC6s6v+9dGsx3cbLxaJ1ul+bQeYu6t4WsT1201VLdc1qjNbTUa1VQYTGQX1yrntXqj1Q5uvr/8PSwX51Kagvwgkh2HAkW3YkSXvRxRsJ3HbzYFjn9qoTBZqDx0R65HKj132LE28zqnMKZp2c2wMF08fO2NXVIqnEHrz1pWi1XqOYjIOTWiyr14FCx+YfKNmP25EJj9UR95FK3o3y4b002bPWhNKGLXPl3QYwSlCYOcXYzGvPWI+TED+QFH7PDM7Q9BcRzfzqHfwmgKpX/aHsWx/2smRuYDh+y5SHDdIaZ7iDA/Yu1xMH/vUlw73zWf/GBqdNJ+uPQmKKaivoyxgW9jjnelRZF8I24rA+qw0KQYD39oUro5NeqmD5ZsirFdwtgkahzq5aNYdzjuAQLkB4oW2hTrYRkBADdHU2K0gxKWehRreyk28G3M8U5RnflE3FYG1KVfKLTT77Ya7/W392anPxahqDsPAKBt1Hufn/aQJfqf/r0u0/78/PzsC5bzybXr2xrIcx0K6036neb721u90bIHEy8UvaVjk7jnUCerbmXCcQ9QnXYoQjjNi5XZTNJMGEbCTM8uHTz0RAi60wEA0JzJZjP7HqdXyPrPZLjZbDZb6LEecllmxkSQE7cUclz/XZxNJxOGkTCTmXzx5HUYRhxjXGB5zBGHqMq4J24rDeryPWJ7za0kQkh99rQfYrgo+W+yuhlQnemybhEUfhvzxHPJRAiZKNyOefSRiY1hcaiVVlUYcdwDVIcHHsu9y4FKrdTieWfI2BDxuwaF+g2xx6dJUKlv93hOOTawPOK4P1CN/oe4zTSoSzaJK6wkqMVCh0XvJqM8HWY/We6JDmpxbciiOy02xg2H3pNq5gccp4IRLIxYH1lQjWWH9ZljbHlTYc1lPZugWjsWrKYZGygMOc6mEu03cetJiOC74Hi/URkk66zJ8pTBY+I630H97Cerl42PfsWhV1NFvsdxtiACPCLuaBHU4yFLbDM2xDQwnljdTATGA2s0Hx9YGHDGawrw0OO8mlHoVVZvJgLY8jje7ymTbLHqiQjwlOUs+8wN46BfeAx6MsLlusScbEAU5jurk44Aj4l9PWWyPdazHoFxz3JLPvMj1okqmOtzRiuhcE9wns1IUh+sjxQAoJlWmf3em2Zzw3BoplXmdse8sgKxr0w79RhUNcJkPog5XoNIMt0wqbWzl5ZK2yH+nymzMAqBMxvnry2VPUG8koo9ZZDvcYZLIXBHcB6NGLUyay8uxVHsT5nCmPWY3GkJiuO4ECvtxGOQpfPSDWKOShBNtsf6OB5RPMeFafZ66VI87WysIGdz+gu8DZdTNaL6ZAlBMW2mppnwKKb3erzwUDC8K42TfCHmsAhxiq2zh9MstoNliBdkOwzqzXFKDudOn0LuVQL++caHWtxwXzC8cwxKVIk5WIKp49m/TPjHE62KDnGDmTaD7FzQ8phzq8fLaTYa9bcI6+9vz7cHswjTbVR/i/zpppJBCMwPYoMVwRAn6Gf8JWZ/ESKb6bFaWdM0Ik0YOkLYr3CjYJH3aBqR6wjcEPuRQKbJoI+M3+KI4V1r0aU7rI80xJ/Xm4mHheHmh6wXHeLN806iwR2XIQ5Q0i+J+TkP0SVbrM/sF+vn4lHVwuX6rIb5lQ6jgfQ7g1ppaWHA8P5oMTBeWKOFLzacU5busF6McOkOq5v5QnQeEW66DPcHAOhnxOzlIQbaJcstfzFnRVmyzWolwyXqrOHcV6pq0UDqnUF1E2Cuz/AuMA6w7XHoGr+Wd4mqzAZruBBO+8NyN7/SRzoi3HQZk03QfnmMXh5isTxhtdLxW2fRx4wq45nlXWqhYNvj0IMRr2yP5WxHBKk3BjWzS30K9s4wHjM2yz3SYrfssMSFoUi7ZNFwDUMVxqxxGWOVbLGonY8IvzsM7/GdmHYO4qHfsGi0p8ct98miybGpBjYFizprGCbVYJG9hnHSr3neSwEjgeQzgy9+YUyg5LBocl3Q4mXc88itlZNK5gY8GpwvGzzcFywanM5jfGDdZRHZR3mMAlYdNZ0sxCX1yiMaWAfl0tL8bC7CDHJgzeERufXL3fXlhdlcLpfBIONvCKLR4+/t1aX52VwuZwBAzuYR9W72NlYK87O5CJNBMx8hyLNvf26uFOZyuVxOhfmgRBxhbGB1HIKIhDsZ2FH+1VnJxzCymAxs27b/6kGwNApDRJ47Hti2/bEKAHggQhCRcEcDO8LuQRDuiRCy5476tm3bKmBloqKVgfgYpyJU5M88WOorCHzm6L9FuECxBQCQvPdCRe39DoJUTUGwksS9AncfYwSpv+JLadujOEDK8pTtSZB7/jqQf4sVFMfhWmmIE6Sv3Xg9aDzQtrpxgPSlEw3kHrx4nXIg/yDiZNyEciqoJtdn7XDAPOjF6j4MYOHeVfPIA+NHW4276Qfps1Gc6IwF6V/9GEG+FcKzEqDWfOZ8zrEAC3968XH3IXxi69FRII6QBzj7q+0p6M0GgF6yRvEZr/BAW77uq+gpwqUXj+NYM6C62A0aH2g8AG3hoGq7sRD3SQUAibWzl4HgiYc0hMbsj9vWxOOND7QgAH355LEnYuGcGyEAtMWf1Y7j8TrrigDSB089x3UdZ/xhbSZAORbOq7VarVa92tBBoZ4r7RydX1v3tUjvjtKgOrGwXjm+uLGqNdk6yIBSTBc3939f3d7V5Or1dx1C6vn1ysnFjVWtRfp32wCVWvbb1s/T69v7mly9KKAyACNfLK0Ui0szGkSKmi9ClIhapAiRo+aLEC2i5ougGrVoESJF1HwR/t///zdSw5Rx6qApG0HaSnmzUinr//YSt41Go/GSnzoLb41Go3GlB5h1VxB10v/2ki0ionFh6nxziYhegpItIqJu5t9equ2zOHVKPs9BqfZ/idH81Fn5n0rRCWH8vLi4OD9I/G/kHxERNV9ERA5qumGapqFjKNSYiApQ46IMiBiwgP4sRE1DAETUEFmoIaKmhUDUfBERAQCh5PoY6KsGNd0wTdPQMRRqTMSY6HO759d/Lataq9VqVcuySn6YWj35+/BSbzQa9ef7852czjDKN9VaYNX6++dwJYlB2tzBlVWtBVcty7J2cdG6tUaS+2jJtyUfzHw/vryx7u5rZ4nslXVXtX5qnFnLsu5qF4kAY2H3/PrWqtZqtdq9ZVk3s/rR7c2zJw3uLPk6ZxxdX1//+WkEYLJ0fFN7qTcajfrT3fmPWZ1hlK+rtcCqdfvnqJTEqLRire8RX+xJOHPSmhBb2Nay7pe4cCjsuHGU9kue9jwKb+F3QSHFHgBg7uLDIf92qjAmInrWOYUxEVE75aOXHgce8d01443Cjgtmg4ionfLBzGFjTGzRuy/qfsbFhMJOmkfpaIyfQwrto621BIUfXiZ9Sg4pFG+LUuJGkEoLN5TgSsuj4OgSv0YU2i1Hht/eBYUfXqV9imNSKF7mosDKhAI9EehsAeDGgPyFMx5PhB8JywQA3CdfT8gBRO8ZAPg2IV8R8hbXHSH8hK+zC7BoU7AQzaj0E5cCPRE4WTNehfD8hHBd1x3ysNQjf+GMRxPXj0Q1CQB4RP5C9vyI3mYimO+R7PWsva2NwO8ZgHyXfPu3m8XC4tLaScuVyD1CADzxaf/Y2NjYKO+cNV0fcYKAxyTbh+UNfgHS3ze2bWlysCF/z0GiSr6T+ulOeWNFj2h5RLLo3lQ2NwLXkri8sXEspNbm2urq6uqKycq1SPY+r8rFwsLS2lHLlcg9RgDd8mlVNuTt03fHR1yq085Jdi9zOvC1U08S1QUdfDG9N5LoIxMkKgi+WqoykqieAP1Ocr4jqDQb0mgeAheGPvb3JIIcjXFHsnOa1SF00ZGedfDnaCeeJO7mdPDFdGUoUWcGQK9K3i/wx+SPsUQDdZmOz50JYTMdkl8zwNR3XckpB7llCNZPPMnOgF71KYHSZCtMWUjjMoJ/NPm+5F0ZoFBRJw2Q6ZD8mgamtuNI7naQ2A8A0E88idQtT6ThIoRenkiTdWCn6hJZqAIWhlJ/FrQriZ5mlKTaIfCQ5NcExGLdlXp5iE03A7A8kSarwE6+SnSvBe0xYH4Y0aYnvZvhykL6SPNwz5PeE0oytjTIA2wJSTQP8iZGpV1K3hnEAg9JftbjtelJH2keVjzp3QT9Plz6Ixo8IvlBD7dL8pPOg2+O1Emr6QZl2xKR6DfuKvMpLZIrSezHQ7d8bjBWeETys86Db47USYN2FS7Vjka79rnDUNqFj4UhCmOpNwN4FAWWbB9ZDFvV3TkTI9qNh/EqeacQK+3S5w5DLIykz9wX0Ks+9+H0e59rCLkwkj6zAJUoABduBsJPFoP6ySxG4W7G5M3nd7z0Wx8rzLzPIP8V7nwsBX99bjHE4ljqzUQFYMxt3XfHnh8RCfvIkJItNeUYNM0vol373IVZ8OnnvgCe+VS1UHjkU9NCFB3pIx0dAOiZwvbtx0j4EDlHCABmI4RuxaZhgv4g0WU8mgH42+dRD1F0pG7mC0DF59UIBdue9G6G2Pak94SSdCeMrKcXy386rkS9PGchqKqonlCCv33utQheglLtANjxpGYyxJYnNU0lyVZEJVeys+FKrjRc4Bk1kh80Jam2NJznyVpmdyyJQ8ZkWd2iTzcTrmkClIXUSkXwnlCw4kjDAs+4J/lRV2I2IprtS+5PDDXbl7wbnYPFkSSOIYKRAgDjRqKqBpB4l8QJBjyEmhtKk/Vw7RTAwkhytlHF4ljqLyrI9STPMliFoeT9hq9gWBJ9rulhjDuJJoeJICw0Se7Px8188HnSAXRLot6q7mO8hcp0JWrOY9CyE2S+SNQtagpyn5L3Oo+h9FuJnF9mEC7USR4ufgn4NpZoaG0s5plZhG9jiZynjVzaNFOZwnmPZO9Wj8rIh1y6dn0sBIAtIdGourmQn00o0K99vO6v5bm8b8UNwrIrUf9mbSHPzCIAGI8SefZFcS6f0zhQHEnkvmzOplNmMlM4tUn2bvWvYVx5EpE7GjDraTB+C4nI6X80m+3uyCPfVh6iKvYG/LFHsrsNAJBt+RC5o0FvXQEU+hKRNxkO5KFDcisJAMmqD5E7GjCfEgAA645E5E2Gg1aOpf8SEpHb/2g32/bII9/mLHwNyNSED7+bATBPJz4hvcYyRlZySWkzLWHp04+IREWFtjn0CV3TAQDyr54Pv2FKxpnrI4/mWZA8G/uEFPVF/CqQPuoJJWCsPQ1FCOfjKIsQ2Tclnl1CCXD5aRy0qwK00utYwWAFfGd+94QiMCsfbqjejATG6tPIC+F8HM2A/DVAy2/ftHr9fr9vB76kQTYXK7edwcQl8tzR58tRaUYD/61+z7Y7q5zko23bdjMPUPjoh7Xt1+NFhMDkyslr17Ztu7Nm1Pr9fq+zygJMrRy/2P1g27a7D+uaH+j5itXq9fv9nh14l/ABzJb/1G25kUs+2rZtv2V8AMzl3dvO0PGIyB33ng5LWQ38ftm2bXe2OYma/VGv1wFWUDggXAgAALArAJ0BKiwBbgA+QRaIRCKhDif0FAICWZu4MF2/AHaAfgB+oH8ATl0QH0/8nPBArr138Zv3B/3fzU1x+pfeH9zP8Jz6tT+ZNxb/iP7H+x/8////1S/sf9s/FL5FeYB+nP9w/l/90/63aN/bz1Cfov/sP8p7zv9o/t3sS/sXqAfzb+L+jB7DPoC/wv+5ekf/r/9v++H0m/tb+wnwKf0X+Z/8/8+/+d9AHoAf//2CP4B6y/TL+AfhL9fvf4bE05pE+6/4ZnNA/JcyJUTgWCA9tESN6mwZptarSiuq3y4hYLjXQzYm68FcAScMDMJiXN+MxHyC5/dpXRQGFa5++aaUzm6q+RVq35FhkJRwXLdNYUPHHR+Dg9EguTRwvO/Fas8xAIPOkSjSdEv0AkNq1O3NpfnkRwBZ4Apce24wLKHvZlawbqcc0W3lSrEoNyAFbX/McnLmKWmmMo0D0hMBXqGTp2CsU3OWeXAA/te/qDNZUG/xg+9Hlbo3m/U0BvWmg8A8nqYduHfTMej1gwngBTTDDauaU2xyT/EKu2yAVo7tz92nx16mCZJP14X59xETg68aGvzjc/O239F2Y3mLhVSkT30J4zGY8qJDBtP/7pb5dprLQRfTR2JsfmxomV3uPx8DSNCrQwoFVLmaUer8yKVWijUvTO4muKJtP6EDgZjCiqq7K3/kXO9cv4c+wGv188VssqGC2U/GtN+XqFvlMc7J+747gdx8+Dii+jhFu7Udv8DCcANDof8sXomK/OpgNUdRoNpYdiAnWsZ8jSlpy199AkcsEf50el8IOCFIiyWdLUZX0OH+Mh1peO9OCKnzVpVlbT4Bf9J5zbSntVCVZz2at+RSqgwwInivhcfxNHkdeXDZRjknloBly9UsAEdQuITzrohxoNC1F9nNG5T2zBkjVC4FduywcLQAdYzS69vomEWmphSW3Z2Yf+7j8ZKGr8GKiSFg87ohLf98v7Ja+tA+j1B1rtUwnOA26ktWlyHl2Uo/KRCrYQJG8DuRyhMQ0WbZKlFYYtqyc2NcZkriN9c2cfxdBHv9pp81DB4GCotRF7XjBL86PS+EHBkTy2FzL/YS8jSUj/1JSRZ3wzhQRgNUlePCV4A+xF3pTnF06Sc6UGWix+GYEGJZCuJJ5odhmvADQs8Sc6aJv1tXbdadu3cP9Nf0dLrzB4Dz2AWtvy3u22nnZxsDQfJaiqqchsWNMXzsI74/h/ECNLMdq37AP/sW40Hhu12xoHWpq8R6nNiflRsu9U0CO0PNptLlZWOeFu3C4UfWSCSG27mdNW9lKmo37Eur0B4W9iuCB+0P/ocMlEkRFXVwB+Jx3kgJV4R9eEcssph69+2HqFBX8wQezQtA2/T+YuLVSH5EW90iWdkfWyQc12kxYNx/EF23SzfrJv1XDXK/aaon6X+6JWHbDWNGTRoBFIO95LMRvtBKvLkm6t394zjXIG/3OrHPAo8hxHOUHfJGbmchEDhgqA4aZqG4g/xG3rrwRibHW0d9b2TQ9LfjXWzM+IIIi7VCh6QBTbH0lZ9fC2B6s1zW0w//Pxs66UU6urx5gJnXG1T3wRn2gT0d1GQggjyZdJopzX9mP4Wo2ebMu+hbpntSAVDgZqzAo+wmPkrUAlbpd/aaDIca9j+qqbfMY+RykhJ4K6un9S571nzBPwUB6dUex51Nx3YxG/BBX/sNpCrX7yG8ylAlcqXihUIilwyLdy4f7mXbtlrzxv1r8wB/fTLI0tqqyI85i6kft3LIZPjHi7SVbbI5Tg2HIvVe250zvNf4wGftNsdkzmDthb///EQIBFqxyBUqAc2pM3At0tIbuGZENklsm0ardy+Y4Vl6N7efg1jfz44wKxTV5lmd2be5o3E/6oYmfbpHG+Fenm7EoLGJ/+SzER7G+2r03wAeNMUcqFjVPIyc8/8DnzhAYlc+V+D9vqU1H8nndHXDbbJiZkO9ybSJbLyhBKV5JT0D8AQyg2ncc1NOHxAJryMC/4wFdc7+Zll3AHgkv13phXhScAyqgF7skn0fU/l+YiJnDncGiqPf7K6SiFjPAUoyi2e/4QCv/rjV8vSA9fTXSNxUa6M4MnBRiYWPXMqGICDBssJPaZBhhwQ468ZgJC4pizRd8XQPOimK6I4oI96Slc/+ZcSh6DKA+ykxhn2OdzmKRwII5hRcMOaLdhmhUwuMgIoLBLHDAIL+dLmBmB2Oi1DE08y5S1kM+sn1fXQM8Wsb5RIYWIzFu6RnTyMchlsI1U6cddFmgtb1+mPqanTj8+hwir9JZs9Sialt2+eB6scLzDorNk6N6Id4BqHSdqcAm4n8xX1WmBlYXDJzNtjfnSzDPRlp2/oyspulBDraGEopWt6+Ioij2CQ3/I3ZyxlwUbORHISAddUK3gxmTUnfGpPer7/W47xNg1UJ5YuI/jR0GDlIZwb5bY93q0o+zMR62xNlYGs01FxXM4EhjovrMXUPLgX6nuLmbwJeTf9ps9L7MfnlYbnJasOr2AcdFAo8S6MUKq02ssY0w8O8ScZLXPzPfqi/PPtOuegfSCITkfyLGMHj7sTVA69lILns5Q1pQ2zgLr4fagwunx0jPrvrYxHGnCZzaiLFcY2kG9NIo7LaeI2Xakc95hA0GrwOYhvIHnSkZjGOr8REBu1Xz7trwrMgbNRmCgdCMpmYAkzVWAPUpBkrWaMz7bMSHrL2eAkjMW5UCZLgV72SG4aErUcQ5sep51lx6BtvdBkjf//sIZLnAciz7IaqqjkzV9bwH9pceEr4aGzPJ+OrXPzPf/1blFM2KXB5IUbYAAAAPV8xc8MxfUAAM5sAAAcClYAABumOAAcdLvKgOjLsK2xRTwXUJsu0jCqgAAA="
logo_data_uri = f"data:image/webp;base64,{VIETTEL_LOGO_BASE64}"

# Bảo vệ biến logo để ứng dụng không dừng nếu logo bị thiếu.
if "logo_data_uri" not in globals():
    logo_data_uri = ""

# Sidebar: thương hiệu, tải file và thống kê.
sidebar_logo_html = (
    f'<div class="sidebar-viettel-logo-wrap">'
    f'<img class="sidebar-viettel-logo" src="{logo_data_uri}" '
    f'alt="Viettel Construction"></div>'
    if logo_data_uri else ""
)

with st.sidebar:
    st.markdown(
        sidebar_logo_html,
        unsafe_allow_html=True,
    )

# Đọc dữ liệu sau khi đã dựng tiêu đề sidebar để file uploader nằm đúng vị trí.
df = get_dataframe()

logo_html = (
    f'<img class="viettel-logo" src="{logo_data_uri}" alt="Viettel Construction">'
    if logo_data_uri else ""
)

# Tạm ẩn banner tiêu đề phía trên. Đổi thành True khi cần hiển thị lại.
SHOW_APP_HEADER = False

if SHOW_APP_HEADER:
    st.markdown(
        f"""
        <div class="app-header">
            <div class="app-header-left">
                <div class="app-copy">
                    <h1>BOT TRA CỨU WO</h1>
                    <p>Tra cứu, thống kê và trực quan hóa WO bằng câu hỏi tiếng Việt</p>
                    <div class="author-line">
                        <span>👤</span><b>Hỗ trợ: Trịnh Bá Cao</b><span>—</span>
                        <a href="mailto:caotb@viettel.com.vn">caotb@viettel.com.vn</a>
                    </div>
                </div>
            </div>
            {logo_html}
        </div>
        """,
        unsafe_allow_html=True,
    )

# Thanh chuyển trang dùng liên kết nội bộ để không phụ thuộc sự kiện widget.
requested_page = query_param_values("page")
page_slug = requested_page[0].strip().lower() if requested_page else "dashboard"
active_page = (
    "Trợ Lý Tra Cứu WO"
    if page_slug == "assistant"
    else "Dashboard Quản Lý WO"
)

dashboard_link_class = "main-page-link active" if page_slug != "assistant" else "main-page-link"
assistant_link_class = "main-page-link active" if page_slug == "assistant" else "main-page-link"

with st.container(key="main_page_switcher"):
    st.markdown(
        f"""
        <style>
        .st-key-main_page_switcher {{
            position:relative !important;
            z-index:2147483647 !important;
            isolation:isolate !important;
            pointer-events:auto !important;
            margin:-20px 0 .5rem -12px !important;
        }}
        .main-page-switcher {{
            position:relative !important;
            z-index:2147483647 !important;
            display:flex !important;
            align-items:center !important;
            gap:5px !important;
            pointer-events:auto !important;
        }}
        .main-page-link {{
            position:relative !important;
            z-index:2147483647 !important;
            display:inline-flex !important;
            align-items:center !important;
            justify-content:center !important;
            min-width:180px !important;
            min-height:40px !important;
            box-sizing:border-box !important;
            padding:7px 14px !important;
            border:1px solid #d9dee7 !important;
            border-radius:0px !important;
            background:#ffffff !important;
            color:#202838 !important;
            text-decoration:none !important;
            font-size:.88rem !important;
            font-weight:650 !important;
            cursor:pointer !important;
            pointer-events:auto !important;
            box-shadow:0 2px 8px rgba(20,32,51,.03) !important;
        }}
        .main-page-link:hover {{
            border-color:#e6002d !important;
            color:#e6002d !important;
            background:#fff5f7 !important;
            text-decoration:none !important;
        }}
        .main-page-link.active {{
            border-color:#e6002d !important;
            background:#e6002d !important;
            color:#ffffff !important;
        }}
        @media(max-width:760px) {{
            .main-page-switcher {{ gap:8px !important; }}
            .main-page-link {{
                min-width:0 !important;
                min-height:38px !important;
                padding:6px 10px !important;
                font-size:.80rem !important;
            }}
        }}
        </style>
        <nav class="main-page-switcher" aria-label="Chuyển trang">
            <a class="{dashboard_link_class}" href="?page=dashboard" target="_self">Dashboard</a>
            <a class="{assistant_link_class}" href="?page=assistant" target="_self">Chatbot Tra Cứu WO</a>
        </nav>
        """,
        unsafe_allow_html=True,
    )

with st.sidebar:
    try:
        file_stat = DEFAULT_FILE.stat()
        file_size_mb = file_stat.st_size / (1024 * 1024)
        file_updated = datetime.fromtimestamp(file_stat.st_mtime)
        file_meta = f"{file_updated.strftime('%d/%m/%Y %H:%M')} - {file_size_mb:.1f} MB"
    except OSError:
        file_updated = datetime.now()
        file_meta = "File dữ liệu hiện tại"

    file_svg = '<svg class="file-icon" viewBox="0 0 24 28" fill="none"><path d="M5 1.5h9l5 5v19H5z" stroke="currentColor" stroke-width="1.8"/><path d="M14 1.5v5h5" stroke="currentColor" stroke-width="1.8"/><path d="M1.5 9.5h11v11h-11z" fill="currentColor"/><path d="M4.2 12.1l2 2.9 2-2.9M4.2 17.9l2-2.9 2 2.9" stroke="white" stroke-width="1.4" stroke-linecap="round"/></svg>'
    trash_svg = '<svg class="trash-icon" viewBox="0 0 24 24" fill="none"><path d="M4 7h16M9 7V4h6v3M7 7l1 13h8l1-13M10 11v5M14 11v5" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>'
    cloud_svg = '<svg width="19" height="19" viewBox="0 0 24 24" fill="none"><path d="M7 18H5.5a3.5 3.5 0 0 1-.5-6.96A6 6 0 0 1 16.5 9a4.5 4.5 0 0 1 1 8.89H17M12 12v8M9 15l3-3 3 3" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>'
    file_html = (
        f'<div class="side-divider"></div><div class="side-section-title">Thống kê tổng quan</div>'
    )
    st.markdown(file_html, unsafe_allow_html=True)

    parquet_mtime = PARQUET_FILE.stat().st_mtime if PARQUET_FILE.exists() else 0.0
    sidebar_summary_slot = st.empty()

    icons = {
        "total": '<svg viewBox="0 0 24 24" fill="none"><path d="M7 3h8l4 4v14H7zM15 3v5h5M10 12h6M10 16h6" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>',
        "open": '<svg viewBox="0 0 24 24" fill="none"><path d="M7 3h10a2 2 0 0 1 2 2v14H5V5a2 2 0 0 1 2-2zM9 8h6M9 12h6M9 16h4" stroke-width="1.8" stroke-linecap="round"/></svg>',
        "closed": '<svg viewBox="0 0 24 24" fill="none"><path d="M7 3h10a2 2 0 0 1 2 2v14H5V5a2 2 0 0 1 2-2zM8.5 12l2.2 2.2L15.8 9" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>',
        "today": '<svg viewBox="0 0 24 24" fill="none"><path d="M5 5h14v15H5zM8 3v4M16 3v4M5 9h14M9 13h2M13 13h2M9 16h2" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>',
    }

    def side_stat(icon_key: str, color_class: str, label: str, value: int) -> None:
        html = (f'<div class="stat-card"><div class="stat-icon {color_class}">{icons[icon_key]}</div>'
                f'<div><div class="stat-label">{label}</div><div class="stat-value">{value:,}</div></div></div>')
        st.markdown(html, unsafe_allow_html=True)

    clock_svg = '<svg class="update-icon" viewBox="0 0 24 24" fill="none"><circle cx="12" cy="12" r="8" stroke="currentColor" stroke-width="1.8"/><path d="M12 7v5l3 2" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"/></svg>'
    update_html = (f'<div class="side-divider"></div><div class="update-row">{clock_svg}<div>'
                   f'<div class="update-label">Dữ liệu cập nhật đến:</div>'
                   f'<div class="update-value">{file_updated.strftime("%d/%m/%Y %H:%M")}</div></div></div>')
    st.markdown(update_html, unsafe_allow_html=True)


    if active_page == "Dashboard Quản Lý WO":
        st.markdown('<div class="side-divider"></div><div class="side-section-title">Bộ lọc Dashboard</div>', unsafe_allow_html=True)

    dashboard_date_col = _pick_dashboard_date_column(df) if active_page == "Dashboard Quản Lý WO" else None
    if dashboard_date_col:
        valid_dates = df[dashboard_date_col].dropna()
        if not valid_dates.empty:
            min_date = valid_dates.min().date()
            max_date = valid_dates.max().date()
            default_dates = (min_date, max_date)
            query_start = query_param_values("date_start")
            query_end = query_param_values("date_end")
            if query_start and query_end:
                try:
                    restored_start = date.fromisoformat(query_start[0])
                    restored_end = date.fromisoformat(query_end[0])
                    if (
                        min_date <= restored_start <= max_date
                        and min_date <= restored_end <= max_date
                        and restored_start <= restored_end
                    ):
                        default_dates = (restored_start, restored_end)
                except ValueError:
                    pass
            selected_dates = st.date_input(
                "Thời gian",
                value=default_dates,
                min_value=min_date,
                max_value=max_date,
                key="dashboard_date_range",
            )
        else:
            selected_dates = None
    else:
        selected_dates = None

    system_options = (
        get_distinct_values(COL_SYSTEM, parquet_mtime)
        if active_page == "Dashboard Quản Lý WO" and PARQUET_FILE.exists()
        else []
    )
    selected_systems = st.multiselect(
        "Hệ thống",
        system_options,
        default=[
            value for value in query_param_values("system")
            if value in system_options
        ],
        placeholder="Tất cả",
        key="dashboard_system_filter",
    ) if system_options else []

    coord_group_options = (
        get_distinct_values(COL_COORD_GROUP, parquet_mtime)
        if active_page == "Dashboard Quản Lý WO" and PARQUET_FILE.exists()
        else []
    )
    selected_coord_groups = st.multiselect(
        "Nhóm điều phối",
        coord_group_options,
        default=[
            value for value in query_param_values("coord_group")
            if value in coord_group_options
        ],
        placeholder="Tất cả",
        key="dashboard_coord_group_filter",
    ) if coord_group_options else []

    employee_options = (
        get_distinct_values(COL_EMPLOYEE, parquet_mtime)
        if active_page == "Dashboard Quản Lý WO" and PARQUET_FILE.exists()
        else []
    )
    selected_employees = st.multiselect(
        "FT / Nhân viên",
        employee_options,
        default=[
            value for value in query_param_values("employee")
            if value in employee_options
        ],
        placeholder="Tất cả",
        key="dashboard_employee_filter",
    ) if employee_options else []

    status_options = (
        get_distinct_values(COL_STATUS, parquet_mtime)
        if active_page == "Dashboard Quản Lý WO" and PARQUET_FILE.exists()
        else []
    )
    selected_statuses = st.multiselect(
        "Trạng thái WO",
        status_options,
        default=[
            value for value in query_param_values("status")
            if value in status_options
        ],
        placeholder="Tất cả",
        key="dashboard_status_filter",
    ) if status_options else []


# Tạo điều kiện SQL Dashboard; không sao chép DataFrame 150.000 dòng.
dashboard_where_sql, dashboard_params = build_sql_where(
    date_column=dashboard_date_col,
    selected_dates=selected_dates,
    systems=selected_systems,
    coord_groups=selected_coord_groups,
    employees=selected_employees,
    statuses=selected_statuses,
)

# Cập nhật 4 thẻ tổng quan tại vị trí đã giữ sẵn trong sidebar.
# Dùng cùng điều kiện và tham số với Dashboard để số liệu luôn đồng bộ.
sidebar_where_sql = (
    dashboard_where_sql if active_page == "Dashboard Quản Lý WO" else ""
)
sidebar_params = (
    tuple(dashboard_params) if active_page == "Dashboard Quản Lý WO" else ()
)
sidebar_summary = get_sidebar_summary(
    parquet_mtime,
    sidebar_where_sql,
    sidebar_params,
)
with sidebar_summary_slot.container():
    side_stat("total", "red", "Tổng WO", sidebar_summary["total"])
    side_stat("open", "orange", "WO tồn", sidebar_summary["opened"])
    side_stat("closed", "green", "WO đã đóng", sidebar_summary["closed"])
    side_stat(
        "today",
        "blue",
        "Hôm nay (FT hoàn thành)",
        sidebar_summary["today_finished"],
    )
    side_stat(
        "today",
        "blue",
        "Hôm qua (FT hoàn thành)",
        sidebar_summary["yesterday_finished"],
    )

# Gắn bộ lọc hiện tại vào liên kết của thẻ KPI để một lần tải lại trang
# vẫn khôi phục đúng phạm vi dữ liệu người dùng đã chọn.
dashboard_link_params = []
if isinstance(selected_dates, (tuple, list)) and len(selected_dates) == 2:
    dashboard_link_params.extend([
        ("date_start", selected_dates[0].isoformat()),
        ("date_end", selected_dates[1].isoformat()),
    ])
for query_name, selected_values in [
    ("system", selected_systems),
    ("coord_group", selected_coord_groups),
    ("employee", selected_employees),
    ("status", selected_statuses),
]:
    dashboard_link_params.extend(
        (query_name, str(value)) for value in selected_values
    )
dashboard_filter_query = urlencode(dashboard_link_params)

if active_page == "Dashboard Quản Lý WO":
    # Dashboard chỉ nhận các bảng tổng hợp nhỏ từ DuckDB.
    render_management_dashboard_sql(
        dashboard_where_sql,
        dashboard_params,
        dashboard_filter_query,
    )

else:
    if "messages" not in st.session_state:
        st.session_state.messages = []

    if not st.session_state.messages:
        st.markdown(
            """
            <div class="welcome-card">
                <div class="welcome-icon">🤖</div>
                <h2>Xin chào, tôi có thể giúp gì về WO?</h2>
                <p>Nhập câu hỏi bằng tiếng Việt để tra cứu, thống kê hoặc vẽ biểu đồ từ dữ liệu WoExport.xlsx.</p>
                <div class="sample-grid">
                    <div class="sample-item">📊 Vẽ biểu đồ đóng WO 7 ngày gần nhất</div>
                    <div class="sample-item">👥 Top 10 FT tồn nhiều nhất</div>
                    <div class="sample-item">⏱️ WO quá hạn của binhnt59</div>
                    <div class="sample-item">🧩 Tổng hợp WO theo hệ thống và nhân viên</div>
                    <div class="sample-item">🏢 Tổng hợp WO theo nhóm điều phối</div>
                    <div class="sample-item">👷 WO tồn của từng nhân viên theo nhóm điều phối</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # Chỉ giữ câu hỏi gần nhất. Trước đây toàn bộ lịch sử bị tính lại
    # sau mỗi lần Streamlit rerun nên càng hỏi nhiều bot càng chậm.
    for message in st.session_state.messages[-2:]:
        with st.chat_message(message["role"]):
            if message["role"] == "user":
                st.markdown(message["content"])
            else:
                answer_question(df, message["content"])

    # Nút Help: bấm để mở/đóng danh sách câu hỏi mẫu mà không cần gõ "help".
    if "show_help_panel" not in st.session_state:
        st.session_state["show_help_panel"] = False

    help_button_label = (
        "✕ Đóng hướng dẫn"
        if st.session_state["show_help_panel"]
        else "❓ Help"
    )

    with st.container(key="help_toggle_area"):
        if st.button(
            help_button_label,
            key="toggle_help_button",
            help="Mở danh sách câu hỏi BOT hỗ trợ",
        ):
            st.session_state["show_help_panel"] = (
                not st.session_state["show_help_panel"]
            )
            st.rerun()

    st.markdown(
        """
        <style>
        .st-key-help_toggle_area {
            max-width:1040px;
            margin:0 auto 6px auto !important;
        }
        .st-key-help_toggle_area [data-testid="stButton"] {
            width:max-content !important;
            margin:0 !important;
        }
        .st-key-help_toggle_area button {
            min-height:34px !important;
            padding:5px 14px !important;
            border:1px solid #f0a0ae !important;
            border-radius:10px !important;
            background:#fff5f7 !important;
            color:#e6002d !important;
            font-size:.88rem !important;
            font-weight:800 !important;
            box-shadow:none !important;
        }
        .st-key-help_toggle_area button:hover {
            border-color:#e6002d !important;
            background:#ffe9ee !important;
            color:#c90028 !important;
            transform:none !important;
        }
        .st-key-help_toggle_area button p {
            color:inherit !important;
            margin:0 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if st.session_state["show_help_panel"]:
        render_help_message()

    typed_question = st.chat_input(
        "Bạn muốn hỏi nội dung nào? Có thể bấm Help để xem câu hỏi mẫu"
    )
    clicked_question = st.session_state.pop("pending_question", None)
    question = clicked_question or typed_question

    if question:
        st.session_state.messages = [
            {"role": "user", "content": question},
            {"role": "assistant", "content": question},
        ]
        st.rerun()
